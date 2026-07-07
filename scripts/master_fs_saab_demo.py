"""SAAB-vocabulary representative demo: prove the ONE master handles a second bank's words.

Same engine, same master, but the TB is captioned the SAAB way — "Loans and advances, net" (not
"Financing"), "Special commission income/expense", "Debt securities in issue and term loans", "Due to
banks" — so the LIVE gpt-5.1-2025-11-13 path must map SAAB's different vocabulary into the SAME master
concepts, and the FS renders in SAAB's OWN labels (bank="saab"). It also POPULATES the SAAB-only master
lines the AlJazira demo left empty (Share premium, Goodwill-as-a-separate-line, Proposed dividends) —
the clearest proof the master is a union and each bank renders its own subset.

Illustrative round numbers (NEVER SAAB's real published figures). Retained earnings is the single
balancing figure; every other line is independent, so the balance check ties only if the derive summed
the rest correctly. No-mappings (live AI) is the essential run; a pre-mapped mirror is also produced.

    python scripts/master_fs_saab_demo.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:                                        # noqa: BLE001
    pass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from ai_accountant.master_fs import (ClientMappingStore, MappingRecord, ProvenanceStore,
                                     apply_mapping_decisions, concept_for_label, generate_master_fs,
                                     load_master_store, propose_account_concepts, propose_master_extensions)
from ai_accountant.master_fs.model import MasterConcept
from ai_accountant.reporting.master_fs_export import export_master_fs_excel, export_master_fs_pdf

SEED_ID = "ksa_bank"      # SAAB is the SAME master, rendered with bank="saab" aliases (the union proof)

M = 1_000  # illustrative SAR'000-ish round numbers (NOT SAAB's real filing)
# (account, SAAB-worded label, current, prior, preparer_concept)  — SAAB vocabulary throughout; several
# sub-accounts roll several-to-one into one master concept. Retained earnings is the balancing figure.
ROWS = [
    # ASSETS
    ("1001", "Cash on hand", 800 * M, 700 * M, "Cash and balances with the central bank (SAMA)"),
    ("1002", "Balances with Saudi Central Bank (SAMA)", 4_200 * M, 3_900 * M, "Cash and balances with the central bank (SAMA)"),
    ("1003", "Statutory deposit with SAMA", 2_500 * M, 2_400 * M, "Cash and balances with the central bank (SAMA)"),
    ("1010", "Placements with banks and other financial institutions", 1_800 * M, 1_600 * M, "Due from banks and other financial institutions, net"),
    ("1011", "Nostro balances with correspondent banks", 600 * M, 550 * M, "Due from banks and other financial institutions, net"),
    ("1020", "Investments held at FVOCI — debt", 9_000 * M, 8_200 * M, "Investments, net"),
    ("1021", "Investments at amortised cost", 5_000 * M, 4_600 * M, "Investments, net"),
    ("1022", "FVOCI equity investments", 1_000 * M, 900 * M, "Investments, net"),
    ("1030", "Corporate loans and advances", 30_000 * M, 27_000 * M, "Loans / financing, net"),
    ("1031", "Consumer loans and advances", 14_000 * M, 12_500 * M, "Loans / financing, net"),
    ("1032", "Credit card lending", 2_000 * M, 1_800 * M, "Loans / financing, net"),
    ("1040", "Property and equipment", 1_200 * M, 1_150 * M, "Property, equipment and right-of-use assets, net"),
    ("1041", "Right-of-use assets", 300 * M, None, "Property, equipment and right-of-use assets, net"),
    ("1050", "Goodwill", 600 * M, 600 * M, "Goodwill and other intangibles, net"),          # SAAB-only line
    ("1051", "Other intangible assets", 200 * M, 180 * M, "Goodwill and other intangibles, net"),  # SAAB-only line
    ("1060", "Other assets and prepayments", 700 * M, 650 * M, "Other assets"),
    # LIABILITIES
    ("2001", "Due to banks", 1_500 * M, 1_300 * M, "Due to banks and other financial institutions"),
    ("2002", "Due to Saudi Central Bank (SAMA)", 500 * M, 450 * M, "Due to banks and other financial institutions"),
    ("2010", "Customers' deposits — demand", 28_000 * M, 25_000 * M, "Customers' deposits"),
    ("2011", "Customers' deposits — time", 12_000 * M, 11_000 * M, "Customers' deposits"),
    ("2012", "Customers' deposits — savings", 8_000 * M, 7_500 * M, "Customers' deposits"),
    ("2020", "Debt securities in issue", 3_000 * M, 3_000 * M, "Debt liabilities (sukuk / debt securities / term loans)"),
    ("2021", "Term loans from financial institutions", 1_500 * M, 1_200 * M, "Debt liabilities (sukuk / debt securities / term loans)"),
    ("2030", "Other liabilities and accruals", 1_200 * M, 1_100 * M, "Other liabilities"),
    # EQUITY
    ("3001", "Share capital", 8_000 * M, 8_000 * M, "Share capital"),
    ("3002", "Share premium", 1_500 * M, 1_500 * M, "Share premium"),                        # SAAB-only line
    ("3003", "Statutory reserve", 3_000 * M, 2_700 * M, "Statutory reserve"),
    ("3004", "Other reserves", 500 * M, 400 * M, "Other reserves"),
    ("3005", "Retained earnings", 2_400 * M, 880 * M, "Retained earnings"),                  # BALANCING figure
    ("3006", "Proposed dividends", 800 * M, 700 * M, "Proposed dividends"),                  # SAAB-only line
    ("3007", "Additional Tier 1 Sukuk", 2_000 * M, 2_000 * M, "Additional Tier 1 sukuk"),
    # INCOME STATEMENT
    ("4001", "Special commission income on loans and advances", 3_200 * M, 2_900 * M, "Special commission / financing income"),
    ("4002", "Special commission income on investments", 900 * M, 820 * M, "Special commission / financing income"),
    ("4010", "Special commission expense on customers' deposits", -1_300 * M, -1_150 * M, "Special commission / financing expense"),
    ("4020", "Fee and commission income", 850 * M, 780 * M, "Fee and commission income"),
    ("4030", "Exchange income, net", 260 * M, 230 * M, "Exchange income, net"),
    ("4040", "Salaries and employee related expenses", -1_400 * M, -1_300 * M, "Salaries and employee-related expenses"),
    ("4041", "Depreciation and amortization", -220 * M, -200 * M, "Depreciation and amortisation"),
    ("4042", "General and administrative expenses", -560 * M, -520 * M, "Other general and administrative expenses"),
    ("4050", "Provision for expected credit losses, net", -480 * M, -410 * M, "Impairment charge / provision for expected credit losses, net"),
    # OTHER COMPREHENSIVE INCOME
    ("5001", "Net changes in fair value of FVOCI equity instruments", 120 * M, -80 * M, "Net change in fair value of FVOCI equity instruments"),
    ("5002", "Debt instrument at FVOCI — net changes in fair value", 180 * M, 90 * M, "Net change in fair value of FVOCI debt instruments"),
]
LABEL = {a: lbl for a, lbl, _c, _p, _pre in ROWS}
TBROWS = [{"account": a, "label": lbl, "current": c, "prior": p} for a, lbl, c, p, _pre in ROWS]

# SAAB-only master lines the AlJazira demo left empty — the union proof
SAAB_ONLY = {"Share premium", "Goodwill and other intangibles, net", "Proposed dividends"}
VOCAB = {"1030": "bs_loans (\"Loans and advances\" ≠ \"Financing\")",
         "4001": "pl_sci (\"Special commission income\")",
         "4010": "pl_sce (\"Special commission expense\")",
         "2020": "bs_debt_liab (\"Debt securities in issue and term loans\")",
         "2001": "bs_due_to (\"Due to banks\")",
         "1050": "bs_goodwill (SAAB-only, separate line)",
         "3002": "bs_share_prem (SAAB-only)",
         "3006": "bs_proposed_div (SAAB-only)"}


def _write_tb_xlsx(path, title, premapped):
    wb = Workbook(); ws = wb.active; ws.title = "TB"
    ws.cell(1, 1, title).font = Font(bold=True)
    heads = (["Account", "Level", "Acct type", "Label", "Mapping (master concept)", "Current", "Prior"]
             if premapped else ["Account", "Label", "Current", "Prior"])
    for c, h in enumerate(heads, 1):
        ws.cell(3, c, h).font = Font(bold=True)
    ws.freeze_panes = ws.cell(4, 1)
    r = 4
    for a, lbl, cur, pri, pre in ROWS:
        vals = ([a, "L4", "C", lbl, pre, cur, "" if pri is None else pri] if premapped
                else [a, lbl, cur, "" if pri is None else pri])
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            if isinstance(v, (int, float)):
                cell.alignment = Alignment(horizontal="right")
        r += 1
    for c in range(1, len(heads) + 1):
        ws.column_dimensions[chr(64 + c)].width = 50 if heads[c - 1] in ("Label", "Mapping (master concept)") else 14
    wb.save(path)


def run_nomappings(store):
    print("=== SAAB-WORDED TB → LIVE gpt-5.1-2025-11-13 mapping (auto-applied, marked ai_assumed) ===")
    GENERIC = {"Other assets", "Other liabilities"}
    items = [(a, LABEL[a]) for a, *_ in ROWS]
    proposals = propose_account_concepts(items, store)                 # LIVE stage 1
    mapped, edge = [], []
    for p in proposals:
        unsure = p.line.strip().lower() == "unsure" or not p.is_confident
        soft_other = p.line in GENERIC and p.confidence != "high"
        (edge if (unsure or soft_other) else mapped).append(p)

    print("  -- vocabulary-difference & SAAB-only lines (the point of this demo) --")
    for p in sorted(proposals, key=lambda x: x.code):
        if p.code in VOCAB:
            disp = "UNSURE" if (p.line.strip().lower() == "unsure" or not p.is_confident) else p.line
            print(f"    {p.code} {LABEL[p.code][:46]:<46} ⇒ {disp[:40]:<40} [{p.confidence}]")
            print(f"          expect {VOCAB[p.code]}  | evidence: {p.evidence[:70]}")
    print("  -- all other lines --")
    for p in sorted(proposals, key=lambda x: x.code):
        if p.code not in VOCAB:
            disp = "UNSURE" if (p.line.strip().lower() == "unsure" or not p.is_confident) else p.line
            print(f"    {p.code} {LABEL[p.code][:46]:<46} ⇒ {disp[:40]:<40} [{p.confidence}]"
                  + ("  (→ 2nd opinion)" if p in edge else ""))

    ext = propose_master_extensions([(p.code, LABEL[p.code]) for p in edge], store) if edge else []
    ext_face = {e.account: e for e in ext if e.is_face_line and e.confidence in ("high", "medium")}
    for e in ext:
        print(f"  -- stage2 {e.account} {LABEL[e.account][:34]:<34} → "
              + (f"ADD '{e.canonical}'" if e.account in ext_face else "keep/finding") + f" [{e.confidence}]")

    ms, au = ClientMappingStore(), ProvenanceStore()
    decisions = {p.code: "assume" for p in mapped}                     # AUTO-APPLY confident maps as ai_assumed
    for p in edge:
        if p.code in ext_face:
            continue
        if p.line in GENERIC and p.line.strip().lower() != "unsure":
            decisions[p.code] = "assume"
    apply_mapping_decisions(proposals, store, client_id="saab_nomap", decisions=decisions,
                            mapping_store=ms, audit=au, approver="auto", at="2026-06-12")
    extensions = []
    for acct, e in ext_face.items():
        cid = "x_" + e.canonical.lower().replace(" ", "_")[:40]
        store.add(MasterConcept(concept_id=cid, statement=e.statement, l0_section=e.l0_section,
                                l1_group=None, canonical_concept=e.canonical, label_aliases={},
                                presence="client_added", order=store.next_order(e.statement, e.l0_section),
                                provenance="proposed_unconfirmed"))
        ms.put("saab_nomap", MappingRecord(account=acct, concept_id=cid, client_label=LABEL[acct],
                                           provenance="ai_assumed", confidence=e.confidence))
        extensions.append({"concept_id": cid, "statement": e.statement, "l0_section": e.l0_section,
                           "canonical": e.canonical, "order": store.get(cid).order,
                           "provenance": "proposed_unconfirmed"})
    stored = {"client": "saab_nomap", "model": "gpt-5.1-2025-11-13", "tb": TBROWS, "extensions": extensions,
              "mappings": [{"account": r.account, "concept_id": r.concept_id, "client_label": r.client_label,
                            "provenance": r.provenance, "confidence": r.confidence,
                            "flagged_reason": r.flagged_reason} for r in ms.mapping("saab_nomap").values()]}
    (ROOT / "exports" / "master_fs_saab_nomap.json").write_text(json.dumps(stored, ensure_ascii=False, indent=2), encoding="utf-8")
    a = sum(1 for m in stored["mappings"] if m["provenance"] == "ai_assumed")
    u = sum(1 for m in stored["mappings"] if not m["concept_id"])
    # did the SAAB-only lines map to their EXISTING concepts (not get proposed as new)?
    cmap = {m["account"]: m["concept_id"] for m in stored["mappings"]}
    print(f"  stored: {a} ai_assumed, {len(extensions)} new concept(s), {u} unmapped→findings")
    for acct in ("1050", "3002", "3006"):
        cid = cmap.get(acct); c = store.get(cid) if cid else None
        print(f"  SAAB-only {acct} {LABEL[acct]:<22} → {c.canonical_concept if c else 'UNMAPPED'} "
              + ("(existing master concept ✓)" if c and c.presence != "client_added" else "(NEW — unexpected)"))
    return stored


def run_premapped():
    print("\n=== SAAB PRE-MAPPED TB → straight-through (preparer), NO AI ===")
    store = load_master_store(seed_id=SEED_ID)
    ms = ClientMappingStore()
    miss = 0
    for a, lbl, _c, _p, pre in ROWS:
        concept = concept_for_label(store, pre)
        miss += concept is None
        ms.put("saab_premap", MappingRecord(account=a, concept_id=(concept.concept_id if concept else None),
                                            client_label=lbl, provenance="preparer", approver="preparer",
                                            approved_at="2026-06-12", flagged_reason="" if concept else "no concept"))
    stored = {"client": "saab_premap", "model": "preparer-provided", "tb": TBROWS, "extensions": [],
              "mappings": [{"account": r.account, "concept_id": r.concept_id, "client_label": r.client_label,
                            "provenance": r.provenance, "confidence": "", "flagged_reason": r.flagged_reason}
                           for r in ms.mapping("saab_premap").values()]}
    (ROOT / "exports" / "master_fs_saab_premap.json").write_text(json.dumps(stored, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  preparer-mapped {len(ROWS)} SAAB lines deterministically ({miss} unmatched)")
    return stored


def _safe(fn, *a):
    try:
        fn(*a); return True
    except PermissionError:
        print(f"  [skip write — file open/locked: {a[-1] if a else ''}]"); return False


def main():
    out = ROOT / "exports"; out.mkdir(exist_ok=True)
    nomap = run_nomappings(load_master_store(seed_id=SEED_ID))
    premap = run_premapped()
    _safe(lambda p: _write_tb_xlsx(p, "Representative SAAB-worded TB — RAW (no mappings) · illustrative", False),
          out / "Representative_TB_saab_nomappings.xlsx")
    _safe(lambda p: _write_tb_xlsx(p, "Representative SAAB-worded TB — PRE-MAPPED (preparer) · illustrative", True),
          out / "Representative_TB_saab_premapped.xlsx")
    for tag, stored in (("saab_nomap", nomap), ("saab_premapped", premap)):
        model = generate_master_fs(stored, strategy="replay", seed_id=SEED_ID, client=stored["client"],
                                   bank="saab",                                  # render in SAAB's OWN labels
                                   period=("31 December 2024", "31 December 2023")).model
        _safe(lambda p, m=model: export_master_fs_excel(m, str(p)), out / f"Master_FS_{tag}.xlsx")
        _safe(lambda p, m=model: export_master_fs_pdf(m, str(p)), out / f"Master_FS_{tag}.pdf")
        # union proof: SAAB-only lines populated on the rendered face
        bs = {l.label for l in model.statements["balance_sheet"]}
        saab_only_shown = [c for c in ("Share premium", "Goodwill and other intangibles, net", "Proposed dividends")
                           if any(c.split(",")[0].split(" and ")[0] in lbl for lbl in bs)]
        ta = next((l.current for l in model.statements["balance_sheet"] if l.concept_id == "bs_total_assets"), 0)
        tle = next((l.current for l in model.statements["balance_sheet"] if l.concept_id == "bs_total_liab_equity"), 0)
        print(f"\n[{tag}] BS lines: {len(model.statements['balance_sheet'])} | findings: {len(model.findings)} | "
              f"NOT-FINAL: {model.not_final} | Total assets {ta:,.0f} vs L&E {tle:,.0f} "
              f"balances={abs(ta - tle) < 0.5} | SAAB-only shown: {saab_only_shown}")
    print(f"\nartifacts in {out}")


if __name__ == "__main__":
    main()
