"""Re-emit the EXISTING bank demo TB + GL JSON fixtures as .xlsx — a FORMAT conversion, not new data.

Same numbers as the JSON (the source of truth); the .xlsx are presentation copies laid out the way a user
would hand over a TB / GL workbook. After writing, each file is read back and asserted to round-trip to the
source JSON to the unit (every value matches, the GL schedules tie + foot to their TB line, the TB still
balances). If anything differs the EMITTER is the bug — never adjust a number to fit.

    python scripts/emit_bank_demo_xlsx.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:  # noqa: BLE001
    pass

from openpyxl import Workbook, load_workbook                                       # noqa: E402
from openpyxl.styles import Font                                                   # noqa: E402

from ai_accountant.master_fs import load_master_store                              # noqa: E402
from ai_accountant.reporting.master_fs_export import build_master_fs_export        # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
BD = ROOT / "tests" / "fixtures" / "bank_demo"
OUT = ROOT / "exports"
TB_JSON = BD / "bank_synthetic_tb.json"
GL_JSON = BD / "bank_synthetic_gl.json"
TB_XLSX = OUT / "bank_synthetic_tb.xlsx"
GL_XLSX = OUT / "bank_synthetic_gl.xlsx"
_BOLD = Font(bold=True)


# ---- TB ------------------------------------------------------------------------------------------
def emit_tb(tb: dict) -> int:
    """One sheet, real-TB shape: Account | Description | Amount (SAR'000). Description is the seed's own
    caption (presentational; the round-trip keys are account + amount)."""
    store = load_master_store(seed_id="ksa_bank")

    def _desc(cid):                                          # the seed's own caption (presentational only)
        c = store.get(cid)
        return (c.canonical_concept if c and c.canonical_concept else cid)
    wb = Workbook(); ws = wb.active; ws.title = "Trial Balance"
    ws.cell(1, 1, f"Synthetic Bank Trial Balance — expressed in {tb['unit']} (synthetic demo data)").font = _BOLD
    for col, head in ((1, "Account"), (2, "Description"), (3, f"Amount ({tb['unit']})")):
        ws.cell(3, col, head).font = _BOLD
    r = 4
    for row in tb["rows"]:
        ws.cell(r, 1, row["account"])
        ws.cell(r, 2, _desc(row["concept_id"]))
        ws.cell(r, 3, row["current"])
        r += 1
    for c, w in (("A", 18), ("B", 56), ("C", 16)):
        ws.column_dimensions[c].width = w
    OUT.mkdir(exist_ok=True); wb.save(TB_XLSX)
    return len(tb["rows"])


def read_tb(path) -> dict:
    ws = load_workbook(path)["Trial Balance"]
    out = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row and row[0] is not None:
            out[str(row[0])] = row[2]
    return out


# ---- GL ------------------------------------------------------------------------------------------
def emit_gl(gl: dict) -> int:
    """One sheet per note (named by the TB concept it reconciles to), roll-forward shape:
    GL account | Asset class | Role | Line | Ref code | Amount. Per account: an Opening row, one row per
    movement ref, a Closing row."""
    sc = gl["sign_convention"]
    wb = Workbook(); wb.remove(wb.active)
    rows_written = 0
    for note_key, body in gl["notes"].items():
        ws = wb.create_sheet(note_key[:31])
        ws.cell(1, 1, f"General Ledger — roll-forward reconciling to TB line '{note_key}' (SAR'000, synthetic)").font = _BOLD
        ws.cell(2, 1, f"Sign convention: accumulated_contra = {sc['accumulated_contra']}; disposals = {sc['disposals']}").font = Font(italic=True, size=9)
        for col, head in ((1, "GL account"), (2, "Asset class"), (3, "Role"), (4, "Line"), (5, "Ref code"), (6, "Amount")):
            ws.cell(4, col, head).font = _BOLD
        r = 5
        for leaf in body["leaves"]:
            acct = leaf["gl_account"]; cls, role = body["confirmed"][acct]
            ws.cell(r, 1, acct); ws.cell(r, 2, cls); ws.cell(r, 3, role)
            ws.cell(r, 4, "Opening balance"); ws.cell(r, 6, leaf["opening"]); r += 1; rows_written += 1
            for ref, amt in leaf["movement_by_ref"].items():
                ws.cell(r, 1, acct); ws.cell(r, 2, cls); ws.cell(r, 3, role)
                ws.cell(r, 4, "Movement"); ws.cell(r, 5, ref); ws.cell(r, 6, amt); r += 1; rows_written += 1
            ws.cell(r, 1, acct); ws.cell(r, 2, cls); ws.cell(r, 3, role)
            ws.cell(r, 4, "Closing balance"); ws.cell(r, 6, leaf["closing"]); r += 1; rows_written += 1
        for c, w in (("A", 18), ("B", 16), ("C", 20), ("D", 16), ("E", 12), ("F", 12)):
            ws.column_dimensions[c].width = w
    OUT.mkdir(exist_ok=True); wb.save(GL_XLSX)
    return rows_written


def read_gl(path) -> dict:
    wb = load_workbook(path)
    notes = {}
    for name in wb.sheetnames:
        ws = wb[name]
        leaves, confirmed = {}, {}
        for row in ws.iter_rows(min_row=5, values_only=True):
            acct, cls, role, line, ref, amt = (list(row) + [None] * 6)[:6]
            if not acct:
                continue
            lf = leaves.setdefault(acct, {"gl_account": acct, "opening": 0, "closing": 0, "movement_by_ref": {}})
            confirmed[acct] = [cls, role]
            if line == "Opening balance":
                lf["opening"] = amt
            elif line == "Closing balance":
                lf["closing"] = amt
            else:
                lf["movement_by_ref"][ref] = amt
        notes[name] = {"leaves": list(leaves.values()), "confirmed": confirmed}
    return notes


# ---- round-trip checks ---------------------------------------------------------------------------
def check_tb(tb: dict) -> tuple:
    back = read_tb(TB_XLSX)
    src = {row["account"]: row["current"] for row in tb["rows"]}
    match = back == src
    # balance the round-tripped values through the engine (read-only — no report regenerated)
    store = load_master_store(seed_id="ksa_bank")
    stored = {"client": tb["client"], "master_id": "ksa_bank", "extensions": [],
              "tb": [{"account": a, "label": a, "current": float(v), "prior": None} for a, v in back.items()],
              "mappings": [{"account": row["account"], "concept_id": row["concept_id"], "client_label": row["concept_id"],
                            "provenance": "preparer"} for row in tb["rows"]]}
    m = build_master_fs_export(store, stored, bank="synthetic bank demo")
    A = next(l.current for l in m.statements["balance_sheet"] if l.concept_id == "bs_total_assets")
    LE = next(l.current for l in m.statements["balance_sheet"] if l.concept_id == "bs_total_liab_equity")
    balances = round(A - LE, 2) == 0.0 and not [f for f in m.findings if str(f[0]) == "balance check"]
    return match, balances, (A, LE)


def check_gl(gl: dict, tb: dict) -> tuple:
    back = read_gl(GL_XLSX)
    tbval = {row["account"]: row["current"] for row in tb["rows"]}
    all_match, all_tie, all_foot = True, True, True
    for key, body in gl["notes"].items():
        src_leaves = {l["gl_account"]: l for l in body["leaves"]}
        bk = back.get(key, {})
        bk_leaves = {l["gl_account"]: l for l in bk.get("leaves", [])}
        if bk_leaves != src_leaves or bk.get("confirmed") != body["confirmed"]:
            all_match = False
        for l in bk_leaves.values():                                   # each schedule ties
            if round(l["opening"] + sum(l["movement_by_ref"].values()) - l["closing"], 2) != 0:
                all_tie = False
        nbv = round(sum(l["closing"] for l in bk_leaves.values()), 2)   # foots to the TB line
        if round(nbv - tbval.get(key, None), 2) != 0:
            all_foot = False
    return all_match, all_tie, all_foot


def main():
    tb = json.loads(TB_JSON.read_text(encoding="utf-8"))
    gl = json.loads(GL_JSON.read_text(encoding="utf-8"))
    n_tb = emit_tb(tb)
    n_gl = emit_gl(gl)
    tb_match, tb_bal, (A, LE) = check_tb(tb)
    gl_match, gl_tie, gl_foot = check_gl(gl, tb)
    Y = lambda b: "Y" if b else "N"  # noqa: E731

    print("=== bank demo input workbooks re-emitted from JSON (format conversion, not new data) ===")
    print(f"  {TB_XLSX.name}: {n_tb} account rows written")
    print(f"  {GL_XLSX.name}: {n_gl} GL rows written ({len(gl['notes'])} note schedules)")
    print("--- HARD CHECK (read back, assert round-trip to the source JSON) ---")
    print(f"  TB  round-trip values match JSON ......... {Y(tb_match)}")
    print(f"  TB  still balances (assets {A:,.0f} == L+E {LE:,.0f}) ... {Y(tb_bal)}")
    print(f"  GL  round-trip leaves/confirmed match JSON {Y(gl_match)}")
    print(f"  GL  schedules tie (opening+moves=closing) . {Y(gl_tie)}")
    print(f"  GL  closings foot to their TB line ........ {Y(gl_foot)}")
    ok = all((tb_match, tb_bal, gl_match, gl_tie, gl_foot))
    print(f"\n{'ALL ROUND-TRIP CHECKS PASSED' if ok else 'MISMATCH — fix the emitter, do not adjust a number'}")
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
