"""Two-TB master-FS demo: a RAW no-mappings TB (live GPT-5.1 mapping, auto-applied as ai_assumed, with
the propose-addition + findings paths) and a PRE-MAPPED TB (straight-through, preparer, NO AI). Writes
both TBs + both Master_FS document sets, replaying the stored live result (no live call at export time).

DEMO HONESTY: the no-mappings run is UNREVIEWED — every AI mapping is applied so the statements populate
but marked ai_assumed (UNCONFIRMED); none is relabelled human-confirmed (that laundering is rejected).
A per-line human confirm step (ai_assumed → ai_confirmed) is a POST-DEMO follow-up (slice-10 GUI).

    python scripts/master_fs_demo.py
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:                                        # noqa: BLE001
    pass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from ai_accountant.master_fs import (ClientMappingStore, MappingRecord, ProvenanceStore, AuditRecord,
                                     apply_mapping_decisions, concept_for_label, generate_master_fs,
                                     load_master_store, propose_account_concepts, propose_master_extensions)
from ai_accountant.master_fs.model import MasterConcept
from ai_accountant.reporting.master_fs_export import export_master_fs_excel, export_master_fs_pdf

SEED_ID = "ksa_bank"      # Slice A: the seed is chosen EXPLICITLY (the future archetype-detector slot)

M = 1_000  # illustrative figures in SAR'000-ish round numbers (NOT the bank's real filing)
# (account, label, current, prior, preparer_concept)  — many accounts roll several-to-one into one
# master concept. preparer_concept (exact master canonical) is used by the PRE-MAPPED TB only.
ROWS = [
    ("1001", "Cash in hand and ATMs", 1_200 * M, 1_050 * M, "Cash and balances with the central bank (SAMA)"),
    ("1002", "Current account with SAMA", 6_500 * M, 5_400 * M, "Cash and balances with the central bank (SAMA)"),
    ("1003", "Statutory deposit with SAMA", 3_800 * M, 3_600 * M, "Cash and balances with the central bank (SAMA)"),
    ("1010", "Money-market placements with banks", 2_400 * M, 2_100 * M, "Due from banks and other financial institutions, net"),
    ("1011", "Nostro current accounts", 700 * M, 650 * M, "Due from banks and other financial institutions, net"),
    ("1020", "FVOCI debt securities (sukuk)", 14_000 * M, 12_500 * M, "Investments, net"),
    ("1021", "FVIS mutual fund units", 1_800 * M, 1_600 * M, "Investments, net"),
    ("1022", "Amortised-cost government sukuk", 9_000 * M, 8_400 * M, "Investments, net"),
    ("1023", "FVOCI equity investments", 1_200 * M, 1_000 * M, "Investments, net"),
    # Corporate Murabaha is the single illustrative BALANCING figure: every other line is chosen
    # independently, and this one is set so assets = liabilities+equity. The balance check then proves
    # the derive pass summed the OTHER ~40 lines correctly (a dropped line would re-open the gap).
    ("1030", "Corporate Murabaha financing", 54_880 * M, 53_460 * M, "Loans / financing, net"),
    ("1031", "Retail personal financing", 21_000 * M, 19_500 * M, "Loans / financing, net"),
    ("1032", "Credit card receivables", 3_000 * M, 2_700 * M, "Loans / financing, net"),
    ("1040", "Land and buildings", 1_500 * M, 1_500 * M, "Property, equipment and right-of-use assets, net"),
    ("1041", "Right-of-use assets (leases)", 400 * M, None, "Property, equipment and right-of-use assets, net"),
    ("1050", "Customers' liability under acceptances", 2_200 * M, 1_900 * M, "Other assets"),
    ("1060", "Prepayments and sundry receivables", 900 * M, 850 * M, "Other assets"),
    ("2001", "Interbank deposits taken", 3_500 * M, 3_100 * M, "Due to banks and other financial institutions"),
    ("2010", "Demand deposits", 48_000 * M, 44_000 * M, "Customers' deposits"),
    ("2011", "Savings accounts", 19_000 * M, 17_500 * M, "Customers' deposits"),
    ("2012", "Time / term deposits", 22_000 * M, 21_000 * M, "Customers' deposits"),
    ("2020", "Tier 2 sukuk issued", 5_000 * M, 5_000 * M, "Debt liabilities (sukuk / debt securities / term loans)"),
    ("2030", "Accrued expenses and accounts payable", 1_400 * M, 1_300 * M, "Other liabilities"),
    ("2040", "Zakat and income tax payable", 600 * M, 520 * M, "Other liabilities"),
    ("3001", "Issued share capital", 10_000 * M, 10_000 * M, "Share capital"),
    ("3002", "Statutory reserve", 4_200 * M, 3_800 * M, "Statutory reserve"),
    ("3003", "Retained earnings", 7_900 * M, 7_100 * M, "Retained earnings"),
    ("3004", "Additional Tier 1 sukuk", 3_000 * M, 3_000 * M, "Additional Tier 1 sukuk"),
    ("4001", "Income from financing", 4_800 * M, 4_400 * M, "Special commission / financing income"),
    ("4002", "Income from investments", 1_600 * M, 1_500 * M, "Special commission / financing income"),
    ("4010", "Return paid on customer deposits", -1_900 * M, -1_700 * M, "Special commission / financing expense"),
    ("4020", "Banking service fees", 1_300 * M, 1_180 * M, "Fee and commission income"),
    ("4030", "Foreign exchange gains, net", 480 * M, 420 * M, "Exchange income, net"),
    ("4040", "Staff salaries and benefits", -2_100 * M, -1_950 * M, "Salaries and employee-related expenses"),
    ("4041", "Depreciation of PP&E and ROU assets", -350 * M, -320 * M, "Depreciation and amortisation"),
    ("4050", "ECL impairment charge on financing", -700 * M, -600 * M, "Impairment charge / provision for expected credit losses, net"),
    ("5001", "Change in FVOCI debt fair value", 220 * M, -140 * M, "Net change in fair value of FVOCI debt instruments"),
    ("5002", "Change in FVOCI equity fair value", 90 * M, 60 * M, "Net change in fair value of FVOCI equity instruments"),
    ("9001", "Sundry suspense — clearing account", 80 * M, 75 * M, "Other assets"),
    ("9002", "Inter-branch control account", 40 * M, 35 * M, "Other assets"),
]
LABEL = {a: lbl for a, lbl, _c, _p, _pre in ROWS}
TBROWS = [{"account": a, "label": lbl, "current": c, "prior": p} for a, lbl, c, p, _pre in ROWS]


def _slug(s):
    return "x_" + re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")[:40]


def _write_tb_xlsx(path, title, premapped):
    wb = Workbook(); ws = wb.active; ws.title = "TB"
    ws.cell(1, 1, title).font = Font(bold=True)
    heads = (["Account", "Level", "Acct type", "Label", "Mapping (master concept)", "Current", "Prior"]
             if premapped else ["Account", "Label", "Current", "Prior"])
    for c, h in enumerate(heads, 1):
        ws.cell(3, c, h).font = Font(bold=True)
    ws.freeze_panes = ws.cell(4, 1)
    r = 4
    for a, lbl, cur, pri, pre in ROWS:
        if premapped:
            vals = [a, "L4", "C", lbl, pre, cur, "" if pri is None else pri]
        else:
            vals = [a, lbl, cur, "" if pri is None else pri]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            if isinstance(v, (int, float)):
                cell.alignment = Alignment(horizontal="right")
        r += 1
    for c in range(1, len(heads) + 1):
        ws.column_dimensions[chr(64 + c)].width = 46 if heads[c - 1] in ("Label", "Mapping (master concept)") else 14
    wb.save(path)


def run_nomappings(store):
    print("=== NO-MAPPINGS TB → LIVE gpt-5.1-2025-11-13 mapping (auto-applied, marked ai_assumed) ===")
    GENERIC = {"Other assets", "Other liabilities"}                # generic buckets → second-opinion
    items = [(a, LABEL[a]) for a, *_ in ROWS]
    proposals = propose_account_concepts(items, store)             # LIVE stage 1
    mapped, edge = [], []
    for p in proposals:
        unsure = p.line.strip().lower() == "unsure" or not p.is_confident
        soft_other = p.line in GENERIC and p.confidence != "high"   # generic bucket, not high confidence
        (edge if (unsure or soft_other) else mapped).append(p)
    for p in sorted(proposals, key=lambda x: x.code):
        disp = "UNSURE" if (p.line.strip().lower() == "unsure" or not p.is_confident) else p.line
        print(f"  {p.code} {LABEL[p.code][:40]:<40} ⇒ {disp[:38]:<38} [{p.confidence}]"
              + ("  (→ 2nd opinion)" if p in edge else ""))
    ext = propose_master_extensions([(p.code, LABEL[p.code]) for p in edge], store) if edge else []  # LIVE stage 2
    ext_face = {e.account: e for e in ext if e.is_face_line and e.confidence in ("high", "medium")}
    print("  -- stage 2 (extension decision on the edge/generic lines) --")
    for e in ext:
        print(f"     {e.account} {LABEL[e.account][:34]:<34} → "
              + (f"ADD '{e.canonical}' [{e.statement}/{e.l0_section}]" if e.account in ext_face
                 else "NOT a distinct face line → keep/finding") + f"  [{e.confidence}]")

    ms, au = ClientMappingStore(), ProvenanceStore()
    decisions = {p.code: "assume" for p in mapped}                  # AUTO-APPLY confident maps as ai_assumed
    for p in edge:
        if p.code in ext_face:
            continue                                               # becomes a NEW concept below (remapped)
        if p.line in GENERIC and p.line.strip().lower() != "unsure":
            decisions[p.code] = "assume"                           # keep the generic-Other mapping (ai_assumed)
        # else: genuinely unsure (junk) → no decision → unmapped → finding
    apply_mapping_decisions(proposals, store, client_id="bank_nomap", decisions=decisions,
                            mapping_store=ms, audit=au, approver="auto", at="2026-06-11")
    extensions = []
    for acct, e in ext_face.items():                               # AI-proposed master additions, auto-added
        cid = _slug(e.canonical)
        store.add(MasterConcept(concept_id=cid, statement=e.statement, l0_section=e.l0_section,
                                l1_group=None, canonical_concept=e.canonical, label_aliases={},
                                presence="client_added", order=store.next_order(e.statement, e.l0_section),
                                provenance="proposed_unconfirmed"))
        ms.put("bank_nomap", MappingRecord(account=acct, concept_id=cid, client_label=LABEL[acct],
                                           provenance="ai_assumed", confidence=e.confidence))
        extensions.append({"concept_id": cid, "statement": e.statement, "l0_section": e.l0_section,
                           "canonical": e.canonical, "order": store.get(cid).order,
                           "provenance": "proposed_unconfirmed"})
    stored = {"client": "bank_nomap", "model": "gpt-5.1-2025-11-13", "tb": TBROWS, "extensions": extensions,
              "mappings": [{"account": r.account, "concept_id": r.concept_id, "client_label": r.client_label,
                            "provenance": r.provenance, "confidence": r.confidence,
                            "flagged_reason": r.flagged_reason} for r in ms.mapping("bank_nomap").values()]}
    (ROOT / "exports" / "master_fs_nomap.json").write_text(json.dumps(stored, ensure_ascii=False, indent=2), encoding="utf-8")
    a = sum(1 for m in stored["mappings"] if m["provenance"] == "ai_assumed")
    u = sum(1 for m in stored["mappings"] if not m["concept_id"])
    print(f"  stored: {a} ai_assumed, {len(extensions)} new concept(s) added, {u} unmapped→findings")
    return stored


def run_premapped():
    print("\n=== PRE-MAPPED TB → straight-through (preparer), NO AI ===")
    store = load_master_store(seed_id=SEED_ID)
    ms = ClientMappingStore()
    miss = 0
    for a, lbl, _c, _p, pre in ROWS:
        concept = concept_for_label(store, pre)
        if concept is None:
            miss += 1
        ms.put("bank_premap", MappingRecord(account=a, concept_id=(concept.concept_id if concept else None),
                                            client_label=lbl, provenance="preparer", approver="preparer",
                                            approved_at="2026-06-11", flagged_reason="" if concept else "no concept"))
    stored = {"client": "bank_premap", "model": "preparer-provided", "tb": TBROWS, "extensions": [],
              "mappings": [{"account": r.account, "concept_id": r.concept_id, "client_label": r.client_label,
                            "provenance": r.provenance, "confidence": "", "flagged_reason": r.flagged_reason}
                           for r in ms.mapping("bank_premap").values()]}
    (ROOT / "exports" / "master_fs_premap.json").write_text(json.dumps(stored, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  preparer-mapped {len(ROWS)} lines deterministically ({miss} unmatched)")
    return stored


def _safe(fn, *a):
    try:
        fn(*a)
        return True
    except PermissionError:
        print(f"  [skip write — file is open/locked: {a[-1] if a else ''}]")
        return False


def main():
    out = ROOT / "exports"; out.mkdir(exist_ok=True)
    nomap = run_nomappings(load_master_store(seed_id=SEED_ID))     # LIVE first → result stored regardless
    premap = run_premapped()
    _safe(lambda p: _write_tb_xlsx(p, "Representative bank TB — RAW (no mappings) · illustrative", False),
          out / "Representative_TB_nomappings.xlsx")
    _safe(lambda p: _write_tb_xlsx(p, "Representative bank TB — PRE-MAPPED (preparer) · illustrative", True),
          out / "Representative_TB_premapped.xlsx")
    for tag, stored in (("nomap", nomap), ("premapped", premap)):
        model = generate_master_fs(stored, strategy="replay", seed_id=SEED_ID, client=stored["client"],
                                   bank=None, period=("31 December 2024", "31 December 2023")).model
        _safe(lambda p, m=model: export_master_fs_excel(m, str(p)), out / f"Master_FS_{tag}.xlsx")
        _safe(lambda p, m=model: export_master_fs_pdf(m, str(p)), out / f"Master_FS_{tag}.pdf")
        print(f"\n[{tag}] BS lines: {len(model.statements['balance_sheet'])} | provenance: "
              f"{len(model.provenance)} | findings: {len(model.findings)} | extensions: "
              f"{len(model.extensions)} | NOT-FINAL: {model.not_final}")
    print(f"\nartifacts in {out}")


if __name__ == "__main__":
    main()
