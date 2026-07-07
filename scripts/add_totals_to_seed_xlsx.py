"""GATE-0 reconciliation: add the four authored BS spine totals to the human-approved seed xlsx.

The four totals (Total assets / liabilities / equity / liabilities and equity) are authored additions
NOT in the original approved xlsx — the same discipline as the earlier dash-drift fix: the xlsx spec is
updated to include them, at the FOOT of their section, so the JSON↔xlsx sequence still reconciles
through GATE 0 (validate_seed). Idempotent — skips if the rows already exist.

    python scripts/add_totals_to_seed_xlsx.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "seeds" / "Master_FS_Structure_Seed_AlJazira_SAAB.xlsx"
SHEET = "Master — Balance Sheet"
TOTALS = {  # canonical -> the concept_id of the LAST leaf of its section (insert the total just after)
    "Total assets": "Other assets",
    "Total liabilities": "Other liabilities",
    "Total equity": "Additional Tier 1 sukuk",
    "Total liabilities and equity": "__after_total_equity__",
}


def _canon(ws, r):
    return str(ws.cell(r, 3).value or "").strip()


def main() -> None:
    wb = load_workbook(XLSX)
    ws = wb[SHEET]
    existing = {_canon(ws, r) for r in range(5, ws.max_row + 1)}
    if "Total assets" in existing:
        print("totals already present — nothing to do (idempotent)")
        return

    def row_of(canonical):
        for r in range(5, ws.max_row + 1):
            if _canon(ws, r) == canonical:
                return r
        raise KeyError(canonical)

    def put_total(at_row, canonical):
        ws.insert_rows(at_row, 1)
        ws.cell(at_row, 3, canonical).font = Font(bold=True)   # col C canonical (bold, it's a total)
        ws.cell(at_row, 6, "both")                              # col F presence; D/E aliases left blank
        ws.cell(at_row, 1, None)                                # col A blank → a concept row, not a section

    # bottom-up so earlier insertions don't shift later anchor rows
    eq_last = row_of("Additional Tier 1 sukuk")                 # foot of EQUITY
    put_total(eq_last + 1, "Total equity")
    put_total(eq_last + 2, "Total liabilities and equity")
    put_total(row_of("Other liabilities") + 1, "Total liabilities")
    put_total(row_of("Other assets") + 1, "Total assets")

    wb.save(XLSX)
    seq = [_canon(ws, r) for r in range(5, ws.max_row + 1) if _canon(ws, r)]
    print("BS canonical sequence now:")
    for c in seq:
        print(("   * " if c.startswith("Total") else "     ") + c)


if __name__ == "__main__":
    main()
