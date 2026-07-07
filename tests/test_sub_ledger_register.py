"""Slice SL-1c.1 — register enrichment (Investment), engine + render, end-to-end.

Proves: read the register → carry rows + N attributes VERBATIM → tie Σ to the bs_investments face (both years,
flag-on-miss) → render the net-new register-table (PDF + Excel). Plus: coarse-vs-enriched dual-mechanism pick;
the layered AI association (Layer 0 deterministic on the real sheet, Layer 1 AI on a synthetic ambiguous name);
the faithful-carry firewall (attribute VALUES never reach a model — only headers/structure); and the column-
LIST-driven render (a different-columns register renders with no code change — the SL-1c.2 reuse proof).

    python tests/test_sub_ledger_register.py
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl                                                                # noqa: E402
from ai_accountant.tb_ingest.register import (propose_register_amount_cols, propose_register_concept,  # noqa: E402
                                              read_register)
from ai_accountant.master_fs.seed import load_master_store                     # noqa: E402
from ai_accountant.reporting.master_fs_export import (build_master_fs_export,   # noqa: E402
                                                      export_master_fs_excel_bytes, export_master_fs_pdf_bytes)

ROOT = Path(__file__).resolve().parent.parent
REAL = ROOT / "Draft Data for working.xlsx"
# the 3 TB investment lines that sum to the bs_investments face (38,967,880 / 36,406,356)
_TB = [("Held at FVIS", 9000000, 8000000), ("Held at FVOCI", 20000000, 19000000),
       ("Held at Amortised Cost", 9967880, 9406356)]


def _stored(store):
    tb = [{"account": f"I{i}", "label": l, "current": float(c), "prior": float(p), "section": "Assets",
           "maturity_hint": "", "levels": ["Assets", "", "Investments, net", l, ""]}
          for i, (l, c, p) in enumerate(_TB)]
    maps = [{"account": r["account"], "concept_id": "bs_investments", "client_label": r["label"],
             "provenance": "preparer"} for r in tb]
    return {"client": "draft", "master_id": store.master_id, "tb": tb, "mappings": maps, "extensions": []}


def _real_register():
    from ai_accountant.tb_ingest.grid import load_grid
    return read_register(load_grid(str(REAL), sheet="Investment "))


def _synthetic_register():
    """A DIFFERENT-columns register (the SL-1c.2 reuse proof): other attribute headers, ties to the same face."""
    grid = [["Ref", "Counterparty", "Rating", "Amount 2025", "Amount 2024"],
            ["A) GROUP ONE", "", "", "", ""],
            ["X-1", "Acme", "Aa3", "38967880", "36406356"]]
    return read_register(grid)


# ============================================================ read + tie + both years
def test_read_register_ties_both_years_with_attributes_and_sections():
    if not REAL.exists():
        print("[skip] real register absent"); return
    reg = _real_register()
    assert "Carrying" in reg["amount_current_header"] and "2025" in reg["amount_current_header"]
    assert "Carrying" in reg["amount_prior_header"] and "2024" in reg["amount_prior_header"]
    assert len(reg["columns"]) == 13 and len(reg["rows"]) >= 55
    assert set(reg["sections"]) >= {"A) HELD AT FVIS", "B) HELD AT FVOCI", "C) HELD AT AMORTISED COST"}
    assert round(sum(r["current"] for r in reg["rows"])) == 38967880
    assert round(sum(r["prior"] or 0 for r in reg["rows"])) == 36406356
    assert any(r["current"] < 0 for r in reg["rows"])                          # the contra rows carried as negatives


def test_register_note_builds_and_ties_to_face_both_years():
    if not REAL.exists():
        print("[skip]"); return
    store = load_master_store(seed_id="ksa_bank")
    m = build_master_fs_export(store, _stored(store), gl={"bs_investments": _real_register()}, notes_attempted=True)
    r = m.note_results["bs_investments"]["register"]
    assert r["status"] == "BUILT"
    assert abs(r["total"] - r["face_value"]) < 0.01 and abs(r["total_prior"] - r["face_value_prior"]) < 0.01


def test_broken_register_does_not_tie_and_BLOCKS():
    if not REAL.exists():
        print("[skip]"); return
    store = load_master_store(seed_id="ksa_bank")
    reg = _real_register()
    reg["rows"] = reg["rows"][:-1]                                             # drop a row → Σ ≠ face
    m = build_master_fs_export(store, _stored(store), gl={"bs_investments": reg}, notes_attempted=True)
    r = m.note_results["bs_investments"]["register"]
    assert r["status"] == "BLOCKED" and r["findings"]                          # never shown as reconciled


# ============================================================ coarse-vs-enriched (existing register-presence pick)
def test_register_present_writes_enriched_absent_writes_coarse():
    store = load_master_store(seed_id="ksa_bank")
    syn = _synthetic_register()
    enriched = build_master_fs_export(store, _stored(store), gl={"bs_investments": syn}, notes_attempted=True)
    assert "register" in enriched.note_results["bs_investments"]               # register present → enriched writes
    coarse = build_master_fs_export(store, _stored(store), notes_attempted=True)
    assert "static" in coarse.note_results["bs_investments"]                   # absent → the SL-1 coarse note (no double-render)


# ============================================================ render — PDF + Excel, attributes VERBATIM, column-list
def test_render_pdf_and_excel_carry_attributes_verbatim():
    if not REAL.exists():
        print("[skip]"); return
    store = load_master_store(seed_id="ksa_bank")
    m = build_master_fs_export(store, _stored(store), gl={"bs_investments": _real_register()}, notes_attempted=True)
    assert len(export_master_fs_pdf_bytes(m)) > 5000                           # the register page renders
    wb = openpyxl.load_workbook(io.BytesIO(export_master_fs_excel_bytes(m)))
    vals = [str(c.value) for s in wb.sheetnames for row in wb[s].iter_rows() for c in row if c.value]
    assert any("Aa3" == v for v in vals) and any("Stage 1" == v for v in vals)  # regulated facts, VERBATIM
    assert any("Credit Rating" in v for v in vals) and any("agrees to" in v for v in vals)


def test_render_is_column_list_driven_different_columns_work():
    store = load_master_store(seed_id="ksa_bank")
    m = build_master_fs_export(store, _stored(store), gl={"bs_investments": _synthetic_register()}, notes_attempted=True)
    nv = next(v for v in m.note_views if v.columns)
    assert nv.columns == ["Ref", "Counterparty", "Rating"]                     # WHATEVER columns the register had
    assert len(export_master_fs_pdf_bytes(m)) > 3000                           # renders with no code change


# ============================================================ AI association — layered, structure only
def test_association_layer0_deterministic_on_the_real_sheet():
    cands = [("bs_investments", "Investments, net")]
    headers = ["Security ID", "Classification", "Carrying Value 2025"]
    out = propose_register_concept("Investment ", headers, cands, client=None)  # NO client → must resolve at Layer 0
    assert out["concept_id"] == "bs_investments" and out["layer"] == 0


def test_association_layer1_fires_on_a_synthetic_ambiguous_name():
    seen = {}

    class _Fake:
        def complete_json(self, prompt, system=None, max_retries=3):
            seen["prompt"] = prompt
            return {"concept_id": "bs_investments", "evidence": "securities headers"}
    cands = [("bs_investments", "Investments, net")]
    headers = ["Security ID", "Classification", "Carrying Value 2025"]
    out = propose_register_concept("Annex 7", headers, cands, client=_Fake())   # name doesn't match → Layer 1
    assert out["concept_id"] == "bs_investments" and out["layer"] == 1
    assert "Security ID" in seen["prompt"]                                     # headers reached the model …


# ============================================================ FAITHFUL-CARRY firewall — attribute VALUES never sent
def test_attribute_values_never_reach_a_model_any_payload():
    captured = []

    class _Capture:
        def complete_json(self, prompt, system=None, max_retries=3):
            captured.append(prompt)
            return {"concept_id": "bs_investments", "current": None, "prior": None, "evidence": ""}
    reg = _synthetic_register() if not REAL.exists() else _real_register()
    headers = reg["columns"] + [reg["amount_current_header"]]
    # EVERY model payload in the slice: the association + the amount-column proposers
    propose_register_concept("Annex 7", headers, [("bs_investments", "Investments, net")], client=_Capture())
    propose_register_amount_cols(headers, client=_Capture())
    blob = " ".join(captured)
    # NO attribute VALUE and NO amount may appear — only headers/structure
    for r in reg["rows"][:20]:
        for v in r["cells"]:
            if str(v).strip() and not any(str(v) in h for h in headers):       # a value that isn't also a header
                assert str(v) not in blob, f"attribute value {v!r} leaked to the model"
        assert str(int(r["current"])) not in blob or r["current"] == 0          # no amount leaked


if __name__ == "__main__":
    failures = 0
    for name, fn in sorted(globals().items()):
        if name.startswith("test_") and callable(fn):
            try:
                fn()
                print(f"[pass] {name}")
            except AssertionError as exc:
                failures += 1
                print(f"[FAIL] {name}: {str(exc).encode('ascii', 'replace').decode()}")
            except Exception as exc:  # noqa: BLE001
                import traceback
                failures += 1
                print(f"[ERROR] {name}: {exc}")
                traceback.print_exc()
    print(f"\n{'ALL PASSED' if not failures else str(failures) + ' FAILED'}")
    sys.exit(1 if failures else 0)
