"""Slice B2b: read the balance-sheet current/non-current split so the face ties to the published statement.

The firewall is the load-bearing part: a BS split is APPLIED only if it reconciles to the TB's lumped total
in BOTH years (no half-apply); a disagreement is shown unsplit + flagged; a BS-only figure with no TB anchor
is injected but marked uncorroborated AND the balance check re-runs on the post-override totals. The TB is
never mutated — it is the anchor the split is checked against.

    python tests/test_bs_split.py
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import ai_accountant.master_fs                                                          # noqa: E402,F401
from ai_accountant.master_fs.face_split import reconcile_face_split                         # noqa: E402
from ai_accountant.master_fs.seed import load_master_store                             # noqa: E402
from ai_accountant.reporting.master_fs_export import build_master_fs_export            # noqa: E402
from ai_accountant.tb_ingest import (best_bs_sheet, confirm_column_roles, detect_header_row,  # noqa: E402
                                     load_grid, parse_bs_split, propose_column_roles, sheet_names)
from tests.test_maturity_split import _mobily_stored                                    # noqa: E402

_ASSET_PAIRS = ("tc_contract_costs_nc", "tc_contract_assets_nc", "tc_fin_other_assets_nc")
# the published current split (riyal-exact) we must tie to
_PUBLISHED = (("tc_contract_costs_nc", "tc_contract_costs_c", 4278.0, 359940.0),
              ("tc_contract_assets_nc", "tc_contract_assets_c", 89959.0, 1003495.0),
              ("tc_fin_other_assets_nc", "tc_fin_other_assets_c", 304722.0, 696921.0))


def _cur(m, cid):
    return next((l.current for st in m.statements.values() for l in st if l.concept_id == cid), None)


def _line(m, cid):
    return next((l for st in m.statements.values() for l in st if l.concept_id == cid), None)


def _schema(grid):
    h = detect_header_row(grid)
    roles = propose_column_roles(grid[h], grid[h + 1:h + 6], client=None)
    return confirm_column_roles(roles, {c.index: c.role for c in roles})


# ============================================================ the pure 4-branch firewall (no store, no I/O)
def test_branch1_both_years_reconcile_applies():
    r = reconcile_face_split([("nc", "c")], {"nc": 1000.0}, {"nc": 800.0},
                           {"nc": {"current": 300.0, "prior": 240.0}, "c": {"current": 700.0, "prior": 560.0}})
    assert r["override"]["nc"]["current"] == 300.0 and r["override"]["c"]["current"] == 700.0
    assert not r["undetermined"] and not r["uncorroborated"] and not r["findings"]


def test_branch2_one_year_disagrees_whole_pair_flags_no_half_apply():
    # current ties (300+700==1000) but PRIOR does not (240+500 != 800) → the WHOLE pair flags, NOT half-applied
    r = reconcile_face_split([("nc", "c")], {"nc": 1000.0}, {"nc": 800.0},
                           {"nc": {"current": 300.0, "prior": 240.0}, "c": {"current": 700.0, "prior": 500.0}})
    assert not r["override"]                                  # no half-apply — the tied year is NOT applied
    assert "nc" in r["undetermined"] and any(k == "disagree" for _c, k, _m in r["findings"])


def test_branch2_current_disagrees_flags():
    r = reconcile_face_split([("nc", "c")], {"nc": 1000.0}, {"nc": 0.0},
                           {"nc": {"current": 300.0}, "c": {"current": 600.0}})   # 900 != 1000
    assert not r["override"] and "nc" in r["undetermined"]


def test_branch3_no_anchor_injects_and_marks_uncorroborated():
    r = reconcile_face_split([("nc", "c")], {}, {},             # neither side in the TB
                           {"nc": {"current": 300.0, "prior": 240.0}, "c": {"current": 700.0}})
    assert r["override"]["nc"]["current"] == 300.0 and r["override"]["c"]["current"] == 700.0
    assert r["uncorroborated"] == {"nc", "c"} and any(k == "no_anchor" for _c, k, _m in r["findings"])


def test_branch4_no_bs_split_keeps_b2a_flag():
    r = reconcile_face_split([("nc", "c")], {"nc": 1000.0}, {"nc": 800.0}, {})    # no BS split
    assert not r["override"] and "nc" in r["undetermined"]
    assert any(k == "undetermined" for _c, k, _m in r["findings"])


def test_branch4_no_bs_both_sides_already_populated_no_flag():
    r = reconcile_face_split([("nc", "c")], {"nc": 300.0, "c": 700.0}, {}, {})
    assert not r["undetermined"] and not r["override"] and not r["findings"]


# ============================================================ the riyal-tie (the published split, through the engine)
def _bs_for_tie(m0):
    amounts = {}
    for nc, c, nc_cur, c_cur in _PUBLISHED:
        lp = _line(m0, nc).prior or 0.0                      # split the lumped PRIOR proportionally so it reconciles too
        nc_pri = round(lp * nc_cur / (nc_cur + c_cur), 2)
        amounts[nc] = {"current": nc_cur, "prior": nc_pri}
        amounts[c] = {"current": c_cur, "prior": round(lp - nc_pri, 2)}
    return {"source": "BS sheet", "amounts": amounts, "unmapped": []}


def test_riyal_tie_face_matches_the_published_split():
    store, stored = _mobily_stored()
    m0 = build_master_fs_export(store, stored)               # baseline: B2a lumps these three
    assert _cur(m0, "tc_contract_costs_nc") == 364218.0 and _cur(m0, "tc_contract_costs_c") is None
    m = build_master_fs_export(store, {**stored, "face_split": _bs_for_tie(m0), "face_split_confirmed": True})
    for nc, c, nc_cur, c_cur in _PUBLISHED:
        assert _cur(m, nc) == nc_cur and _cur(m, c) == c_cur                  # ties, riyal-exact
    su = {l.concept_id for st in m.statements.values() for l in st if l.split_undetermined}
    assert not (set(_ASSET_PAIRS) & su)                      # split_undetermined CLEARED
    assert not any(str(f[0]).startswith("maturity:") and "contract" in str(f[0]) for f in m.findings)
    # branch-1 is BALANCE-NEUTRAL (Σ == the anchor it replaces): the balance state is unchanged vs the baseline
    # (_mobily_stored is pre-close, so the baseline itself carries the unclosed-retained balance finding)
    bc = lambda mm: any(f[0] == "balance check" for f in mm.findings)
    assert bc(m) == bc(m0)


# ============================================================ reconcile-fail, both-years guard, pending
def test_reconcile_fail_flags_keeps_lumped_not_applied():
    store, stored = _mobily_stored()
    bs = {"source": "BS sheet", "unmapped": [],
          "amounts": {"tc_contract_costs_nc": {"current": 4278.0, "prior": 0.0},
                      "tc_contract_costs_c": {"current": 999.0, "prior": 0.0}}}   # 5,277 != 364,218
    m = build_master_fs_export(store, {**stored, "face_split": bs, "face_split_confirmed": True})
    nc = _line(m, "tc_contract_costs_nc")
    assert nc.current == 364218.0 and nc.split_undetermined and _cur(m, "tc_contract_costs_c") is None
    assert m.not_final and any("disagree" in str(f[2]).lower() for f in m.findings if "contract_costs" in str(f[0]))


def test_both_years_guard_current_ties_prior_doesnt_flags_whole_pair():
    store, stored = _mobily_stored()
    m0 = build_master_fs_export(store, stored)
    lp = _line(m0, "tc_contract_costs_nc").prior
    bs = {"source": "BS sheet", "unmapped": [],
          "amounts": {"tc_contract_costs_nc": {"current": 4278.0, "prior": 1.0},      # current ties...
                      "tc_contract_costs_c": {"current": 359940.0, "prior": lp}}}      # ...prior sums to lp+1 ≠ lp
    m = build_master_fs_export(store, {**stored, "face_split": bs, "face_split_confirmed": True})
    nc = _line(m, "tc_contract_costs_nc")
    assert nc.current == 364218.0 and nc.split_undetermined          # NOT half-applied on the tied year
    assert _cur(m, "tc_contract_costs_c") is None


def test_unconfirmed_bs_split_is_pending_not_final_not_applied():
    store, stored = _mobily_stored()
    bs = {"source": "BS sheet", "unmapped": [],
          "amounts": {"tc_contract_costs_nc": {"current": 4278.0}, "tc_contract_costs_c": {"current": 359940.0}}}
    m = build_master_fs_export(store, {**stored, "face_split": bs})    # read but NOT confirmed
    assert _cur(m, "tc_contract_costs_nc") == 364218.0              # not applied
    assert m.not_final and any(f[0] == "face_split:pending" for f in m.findings)


def test_no_bs_split_keeps_b2a_unchanged():
    store, stored = _mobily_stored()
    m = build_master_fs_export(store, stored)
    su = {l.concept_id for st in m.statements.values() for l in st if l.split_undetermined}
    assert set(_ASSET_PAIRS) <= su
    assert len([f for f in m.findings if str(f[0]).startswith("maturity:")]) == 3


# ============================================================ no-anchor injection + the balance guardrail
def _minimal_balanced_stored(store):
    # one asset == one equity → the SOFP balances WITHOUT any contract-costs money in the TB
    def m(acct, cid, label):
        return {"account": acct, "concept_id": cid, "client_label": label, "provenance": "preparer",
                "confidence": "", "flagged_reason": ""}
    return {"client": "x", "master_id": store.master_id, "extensions": [],
            "tb": [{"account": "a1", "label": "Cash", "current": 1000.0, "prior": 1000.0},
                   {"account": "a2", "label": "Share capital", "current": 1000.0, "prior": 1000.0}],
            "mappings": [m("a1", "tc_cash", "Cash"), m("a2", "tc_share_cap", "Share capital")]}


def test_no_anchor_injects_marked_uncorroborated_and_balance_guardrail_fires():
    store = load_master_store(seed_id="telecom_mobily")
    base = _minimal_balanced_stored(store)
    m0 = build_master_fs_export(store, base)
    assert not any(f[0] == "balance check" for f in m0.findings)    # balances WITHOUT the BS-only money
    bs = {"source": "BS sheet", "unmapped": [],                     # contract costs are NOT in this TB → no anchor
          "amounts": {"tc_contract_costs_nc": {"current": 4278.0, "prior": 4278.0},
                      "tc_contract_costs_c": {"current": 359940.0, "prior": 359940.0}}}
    m = build_master_fs_export(store, {**base, "face_split": bs, "face_split_confirmed": True})
    cc = _line(m, "tc_contract_costs_c")
    assert cc is not None and cc.current == 359940.0 and cc.uncorroborated   # injected + marked, never "tied"
    assert any(f[0] == "balance check" for f in m.findings)         # the injection unbalanced the SOFP → flagged loud
    assert m.not_final


# ============================================================ reading the BS sheet (ingest reuse, anti-hardcoding)
def test_parse_bs_split_resolves_labels_to_nc_c_concepts_and_flags_unmapped():
    store = load_master_store(seed_id="telecom_mobily")
    grid = [["Particulars", "2024", "2023"],
            ["Contract costs - non-current", 4278, 4000],     # hyphen (resolve normalises dash variants)
            ["Contract costs - current", 359940, 350000],
            ["Mystery line not in the seed", 111, 222]]
    bs = parse_bs_split(grid, _schema(grid), store, client=None)
    assert bs["amounts"]["tc_contract_costs_nc"]["current"] == 4278.0
    assert bs["amounts"]["tc_contract_costs_c"]["current"] == 359940.0
    assert any("Mystery" in u["label"] for u in bs["unmapped"])    # unmapped → flagged, not mis-assigned


def test_best_bs_sheet_scored_not_a_literal_and_excludes_the_tb():
    assert best_bs_sheet(["TB Mapping", "Balance Sheet"]) == "Balance Sheet"
    assert best_bs_sheet(["Sheet1", "Statement of Financial Position"]) == "Statement of Financial Position"
    assert best_bs_sheet(["TB", "SOFP"], exclude="TB") == "SOFP"
    assert best_bs_sheet(["GL", "BS 2024"]) == "BS 2024"           # non-Mobily name, still found
    assert best_bs_sheet(["TB Mapping"]) is None                   # no BS-ish name → no guess (B2a stays)


def test_multisheet_workbook_bs_sheet_read_from_bytes_non_mobily_layout():
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    tb = wb.create_sheet("TB Mapping")
    tb.append(["Account", "L1", "2024"])
    tb.append(["Contract costs", "Non-Current Assets", 364218])
    bs = wb.create_sheet("Statement of Financial Position")        # non-'BS' sheet name
    bs.append(["2024", "Description"])                             # amount-FIRST layout (structure-agnostic)
    bs.append([4278, "Contract costs - non-current"])
    bs.append([359940, "Contract costs - current"])
    buf = io.BytesIO()
    wb.save(buf)
    names = sheet_names(io.BytesIO(buf.getvalue()))
    chosen = best_bs_sheet(names, exclude="TB Mapping")
    assert chosen == "Statement of Financial Position"
    grid = load_grid(io.BytesIO(buf.getvalue()), filename="x.xlsx", sheet=chosen)
    store = load_master_store(seed_id="telecom_mobily")
    split = parse_bs_split(grid, _schema(grid), store, client=None)
    assert split["amounts"]["tc_contract_costs_nc"]["current"] == 4278.0   # resolved despite the odd layout/name
    assert split["amounts"]["tc_contract_costs_c"]["current"] == 359940.0


# ============================================================ B2c: LIVE AI-fallback for foreign BS labels
class _Fake:
    """A one-off proposer (NOT _DemoLLM — keeps the fence clean): maps a label substring to a canonical concept
    by reading the prompt's 'account<tab>label' rows. `mapping={}` → every line is 'unsure'."""
    def __init__(self, mapping=None):
        self.mapping = mapping or {}
        self.called = False

    def complete_json(self, prompt, system=None, **k):
        self.called = True
        props = []
        for ln in str(prompt).splitlines():
            if "\t" in ln:                                   # only the data rows are 'acct<TAB>label'
                acct, _, label = ln.partition("\t")
                for sub, canon in self.mapping.items():
                    if sub.lower() in label.lower():
                        props.append({"account": acct.strip(), "concept": canon,
                                      "confidence": "high", "evidence": "test"})
                        break
        return {"proposals": props}


class _Raise:
    """A client that must NEVER be called — proves the deterministic path makes no model call."""
    def complete_json(self, *a, **k):
        raise AssertionError("the live model was called for a CLEAN-label BS line (should be deterministic)")


def _bs_grid(rows):
    return [["Particulars", "2024", "2023"], *rows]


def test_foreign_label_routes_to_live_ai_fallback_and_resolves():
    store = load_master_store(seed_id="telecom_mobily")
    grid = _bs_grid([["Long-term other financial assets", 304722, 241000],     # FOREIGN — no '— non-current' suffix
                     ["Short-term other financial assets", 696921, 552000]])
    fake = _Fake({"long-term other financial": "Financial and other assets — non-current",
                  "short-term other financial": "Financial and other assets — current"})
    bs = parse_bs_split(grid, _schema(grid), store, client=fake)
    assert fake.called                                       # the AI-fallback FIRED (deterministic couldn't place)
    assert bs["amounts"]["tc_fin_other_assets_nc"]["current"] == 304722.0
    assert bs["amounts"]["tc_fin_other_assets_c"]["current"] == 696921.0
    assert {a["concept_id"] for a in bs["ai_proposed"]} == {"tc_fin_other_assets_nc", "tc_fin_other_assets_c"}
    assert not bs["unmapped"]


def test_clean_labels_resolve_deterministically_no_model_call():
    store = load_master_store(seed_id="telecom_mobily")
    grid = _bs_grid([["Contract costs - non-current", 4278, 4000],             # clean suffix → resolve_row places it
                     ["Contract costs - current", 359940, 350000]])
    bs = parse_bs_split(grid, _schema(grid), store, client=_Raise())           # _Raise fires if the model is touched
    assert bs["amounts"]["tc_contract_costs_nc"]["current"] == 4278.0          # resolved with NO model call
    assert not bs["ai_proposed"] and not bs["unmapped"]


def test_ai_unsure_flags_never_guesses():
    store = load_master_store(seed_id="telecom_mobily")
    grid = _bs_grid([["Some utterly foreign line", 1, 2]])
    bs = parse_bs_split(grid, _schema(grid), store, client=_Fake({}))          # fake returns 'unsure' for all
    assert not bs["amounts"] and not bs["ai_proposed"]
    assert bs["unmapped"] and "unsure" in bs["unmapped"][0]["reason"]


def test_ai_offpair_concept_is_flagged_not_a_silent_noop():
    store = load_master_store(seed_id="telecom_mobily")
    grid = _bs_grid([["Petty cash float", 100, 90]])
    # the AI guesses a NON-pair concept (Cash) — decision H: that is off-task → FLAG, never silently mapped
    bs = parse_bs_split(grid, _schema(grid), store, client=_Fake({"petty cash": "Cash and cash equivalents"}))
    assert "tc_cash" not in bs["amounts"] and not bs["ai_proposed"]
    assert bs["unmapped"] and "off-pair" in bs["unmapped"][0]["reason"]


def test_ai_proposed_concept_whose_amounts_dont_reconcile_still_flags():
    store, stored = _mobily_stored()
    grid = _bs_grid([["Long-term other financial assets", 304722, 241000],
                     ["Short-term other financial assets", 999, 1]])           # nc+c != the TB lump (1,001,643)
    fake = _Fake({"long-term other financial": "Financial and other assets — non-current",
                  "short-term other financial": "Financial and other assets — current"})
    bs = parse_bs_split(grid, _schema(grid), store, client=fake)
    assert bs["ai_proposed"]                                  # the AI proposed the concepts...
    m = build_master_fs_export(store, {**stored, "face_split": bs, "face_split_confirmed": True})
    nc = _line(m, "tc_fin_other_assets_nc")
    assert nc.current == 1001643.0 and nc.split_undetermined  # ...but the reconcile firewall STILL flags the amounts
    assert any("disagree" in str(f[2]).lower() for f in m.findings if "fin_other_assets" in str(f[0]))


def test_live_ai_fallback_uses_the_real_client_if_key_present():
    import os
    if not os.environ.get("OPENAI_API_KEY"):
        print("[skip] no OPENAI_API_KEY — the live BS-split AI fallback is not exercised")
        return
    from ai_accountant.llm.client import LLMClient
    store = load_master_store(seed_id="telecom_mobily")
    grid = _bs_grid([["Long-term other financial assets", 304722, 241000]])
    bs = parse_bs_split(grid, _schema(grid), store, client=LLMClient())        # LIVE model
    # the live model either places it on a maturity-half concept or it is flagged — never silently dropped
    assert (bs["ai_proposed"] or bs["unmapped"])


# ============================================================ the two B2b gaps closed here
def test_no_bs_path_is_byte_identical_frozen_replay():
    import json
    base = Path(__file__).resolve().parent / "fixtures" / "regression" / "mobily_nobs_face.json"
    store, stored = _mobily_stored()
    m = build_master_fs_export(store, stored)                 # the TB-only path (no face_split)
    face = [[l.concept_id, l.current, l.prior, l.kind, l.derived, l.split_undetermined,
             getattr(l, "uncorroborated", False)] for st in m.statements.values() for l in st]
    snap = {"face": face, "finding_keys": sorted(str(f[0]) for f in m.findings)}
    saved = json.loads(base.read_text(encoding="utf-8"))
    assert snap == saved, "the no-BS (B2a) path changed — frozen replay broke"


def test_unit_mismatch_flags_never_silently_scaled():
    store, stored = _mobily_stored()
    m0 = build_master_fs_export(store, stored)
    # a BS sheet quoted ×1000 vs the TB → nc+c is 1000× the lump → must FLAG (branch 2), never auto-scaled to fit
    bs = {"source": "BS sheet", "ai_proposed": [], "unmapped": [],
          "amounts": {"tc_contract_costs_nc": {"current": 4278000.0, "prior": (m0 and 0.0)},
                      "tc_contract_costs_c": {"current": 359940000.0, "prior": 0.0}}}
    m = build_master_fs_export(store, {**stored, "face_split": bs, "face_split_confirmed": True})
    nc = _line(m, "tc_contract_costs_nc")
    assert nc.current == 364218.0 and nc.split_undetermined   # lumped retained — NOT scaled to reconcile
    assert any("disagree" in str(f[2]).lower() for f in m.findings if "contract_costs" in str(f[0]))


if __name__ == "__main__":
    failures = 0
    for name, fn in sorted(globals().items()):
        if name.startswith("test_") and callable(fn):
            try:
                fn()
                print(f"[pass] {name}")
            except AssertionError as exc:
                failures += 1
                print(f"[FAIL] {name}: {exc}")
            except Exception as exc:  # noqa: BLE001
                import traceback
                failures += 1
                print(f"[ERROR] {name}: {exc}")
                traceback.print_exc()
    print(f"\n{'ALL PASSED' if not failures else str(failures) + ' FAILED'}")
    sys.exit(1 if failures else 0)
