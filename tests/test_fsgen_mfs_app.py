"""Slice G1 UI regression: the master-FS (seed-driven) view drives the engine's confirm chain faithfully.

Drives `fsgen_app.py` via AppTest (the app boots straight into the seed-driven flow — the legacy GL view and
its toggle were removed) and asserts each of the six gates is a REAL stop, and the five auto-accept risks each
force a human step:
  1. archetype gate has NO default pick; the ambiguous fixture BLOCKS on 'unsure' (risk #1, #5)
  2. mapping is PER LINE — a line left unreviewed stays AI_ASSUMED (amber) and keeps not_final True (risk #2)
  3. the split renders JUDGMENT_UNCONFIRMED until an explicit confirm, then JUDGMENT_CONFIRMED with the
     persistent marker (risk #3) — building never auto-confirms
  4. figures are WITHHELD until the unit is set; export is enabled only after (risk #4)
  5. 'type your own' edits a LABEL only — there is NO monetary-amount box on any reconciling line
Plus the FENCE: `_DemoLLM` (the offline stub) is structurally confined to the demo module — the
ai_accountant engine package never imports it, and the production proposers default to the real LLMClient.

    python tests/test_fsgen_mfs_app.py
"""
from __future__ import annotations

import inspect
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

ROOT = Path(__file__).resolve().parent.parent
APP = ROOT / "fsgen_app.py"
_CONFIRM = "✓ Confirm AI proposal"


def _surface(at):
    out = []
    for attr in ("markdown", "caption", "error", "warning", "info", "subheader"):
        out += [str(w.value) for w in getattr(at, attr, [])]
    return " || ".join(out)


def _ss(at, key):
    return at.session_state[key] if key in at.session_state else None


def _open_view(case):
    from streamlit.testing.v1 import AppTest
    at = AppTest.from_file(str(APP), default_timeout=300).run()
    at.radio(key="mfs_dataset").set_value(case).run()
    at.button(key="mfs_load").click().run()
    return at


def _confirm_archetype(at, seed_id):
    at.radio(key="mfs_pick").set_value(seed_id).run()
    at.button(key="mfs_confirm_arch").click().run()
    return at


def _apply_mappings(at, leave_last_assumed=True):
    sels = [s for s in at.selectbox if str(getattr(s, "key", "")).startswith("mfs_map_")]
    confirmable = sels[:-1] if leave_last_assumed else sels
    for s in confirmable:
        s.set_value(_CONFIRM)
    at.run()
    at.button(key="mfs_apply_map").click().run()
    return at


# ----------------------------------------------------------------------------- bank end-to-end
def test_bank_walkthrough_gates_block_and_engine_drives_honesty():
    at = _open_view("bank")
    assert not at.exception, at.exception
    # G1 — NO default archetype pick; the gate stops, nothing downstream
    assert at.radio(key="mfs_pick").value is None
    assert _ss(at, "mfs_seed_id") is None
    assert "Map account" not in " ".join(h.value for h in at.header)        # mapping gate not reached
    _confirm_archetype(at, "ksa_bank")
    assert _ss(at, "mfs_seed_id") == "ksa_bank"

    # G2 — per-line mapping; leave the LAST line unreviewed → it stays AI_ASSUMED / not_final
    assert len([s for s in at.selectbox if str(getattr(s, "key", "")).startswith("mfs_map_")]) > 10
    assert len(at.number_input) == 0                                        # NO monetary-amount box anywhere
    _apply_mappings(at, leave_last_assumed=True)
    assert _ss(at, "mfs_mapping_confirmed") is True
    model = _ss(at, "mfs_model")
    assert any(l.ai_assumed for st in model.statements.values() for l in st)   # unreviewed line stayed AI_ASSUMED
    assert model.not_final is True                                          # engine-derived, cannot be cleared
    assert "AI assumption" in _surface(at)

    # G3/G4 — figures WITHHELD until the unit; note statuses read in WORDS; export gated on the unit
    surf = _surface(at)
    assert "WITHHELD" in surf and "NOT FINAL" in surf
    assert "BUILT" in surf and "not generated" in surf                      # note status words, not just colour
    assert len(at.get("download_button")) == 0                              # no export before the unit
    at.selectbox(key="mfs_unit").set_value("SAR'000").run()
    assert len(at.get("download_button")) == 2                              # both downloads appear after the unit
    assert not at.exception, at.exception

    # export returns non-empty bytes through the same model the screen rendered
    from ai_accountant.reporting.master_fs_export import (export_master_fs_excel_bytes,
                                                          export_master_fs_pdf_bytes)
    m = _ss(at, "mfs_model")
    assert len(export_master_fs_excel_bytes(m)) > 0 and len(export_master_fs_pdf_bytes(m)) > 0


# ----------------------------------------------------------------------------- telecom split (judgment) gate
def test_telecom_split_is_judgment_unconfirmed_until_an_explicit_confirm():
    at = _open_view("telecom")
    _confirm_archetype(at, "telecom_mobily")
    assert _ss(at, "mfs_seed_id") == "telecom_mobily"
    _apply_mappings(at, leave_last_assumed=False)
    at.selectbox(key="mfs_unit").set_value("SAR'000").run()


    def _split(at):
        return next(v for v in _ss(at, "mfs_model").note_views if v.note_ref == "lease_liabilities")

    # the SOFP balances and the P&L carries into OCI (a real result, the engine's hard gate)
    m0 = _ss(at, "mfs_model")
    assert not [f for f in m0.findings if str(f[0]) == "balance check"]
    # BY DEFAULT the split is a management judgment, UNCONFIRMED — building never auto-confirmed it
    assert _split(at).status == "JUDGMENT_UNCONFIRMED"
    assert "JUDGMENT" in _surface(at) and "UNCONFIRMED" in _surface(at)
    # an EXPLICIT confirm flips it; the persistent marker survives (Case B can never read as reconciled)
    at.button(key="mfs_confirm_split").click().run()
    assert _ss(at, "mfs_split_confirmed") is True
    m1 = _ss(at, "mfs_model")
    assert _split(at).status == "JUDGMENT_CONFIRMED"
    assert m1.judgment_markers                                              # marker persists even after confirm
    assert len(at.get("download_button")) == 2


# ----------------------------------------------------------------------------- ambiguous → unsure block
def test_ambiguous_blocks_with_no_default_until_a_seed_is_chosen():
    at = _open_view("ambiguous")
    assert "UNSURE" in _surface(at)                                         # the detector lands on unsure
    assert at.radio(key="mfs_pick").value is None                          # NO seed pre-selected
    assert _ss(at, "mfs_seed_id") is None                                  # nothing proceeds
    assert "Map account" not in " ".join(h.value for h in at.header)
    # the human MUST pick explicitly to move on (a preparer override after the block)
    _confirm_archetype(at, "ksa_bank")
    assert _ss(at, "mfs_seed_id") == "ksa_bank"


# ----------------------------------------------------------------------------- type-your-own edits a LABEL only
def test_override_edits_a_concept_label_never_an_amount():
    at = _open_view("bank")
    _confirm_archetype(at, "ksa_bank")
    sels = [s for s in at.selectbox if str(getattr(s, "key", "")).startswith("mfs_map_")]
    # every override option is a concept LABEL (canonical), and there is NO numeric input on the page
    sample = sels[0]
    overrides = [o for o in sample.options if str(o).startswith("Override → ")]
    assert overrides and all(not o[len("Override → "):].strip().lstrip("-").isdigit() for o in overrides)
    assert len(at.number_input) == 0
    # choosing an override maps a line onto the chosen CONCEPT (a label), computed by the engine
    target = overrides[0][len("Override → "):]
    sample.set_value(overrides[0]).run()
    for s in sels[1:]:
        s.set_value(_CONFIRM)
    at.run()
    at.button(key="mfs_apply_map").click().run()
    stored = _ss(at, "mfs_stored")
    from ai_accountant.master_fs.seed import load_master_store
    store = load_master_store(seed_id="ksa_bank")
    target_id = store.by_canonical(target).concept_id
    assert target_id in {m["concept_id"] for m in stored["mappings"]}       # the override placed a line on the chosen concept


# ----------------------------------------------------------------------------- the _DemoLLM fence
def test_demo_llm_is_structurally_confined_to_the_demo_module():
    import fsgen_mfs
    # (a) the offline stub cannot be constructed without a demo fixture — it can never reach real client data
    try:
        fsgen_mfs._DemoLLM(object())
        raise AssertionError("_DemoLLM accepted a non-fixture")
    except TypeError:
        pass
    # (b) the ai_accountant ENGINE package never imports the stub or the demo module (grep-clean)
    pkg = ROOT / "ai_accountant"
    hits = [p.name for p in pkg.rglob("*.py")
            if "_DemoLLM" in p.read_text(encoding="utf-8") or "fsgen_mfs" in p.read_text(encoding="utf-8")]
    assert not hits, f"engine package references the demo stub: {hits}"
    # (c) the PRODUCTION proposers default to the real LLMClient (no client arg → real model, not a stub)
    from ai_accountant.master_fs.mapping import (propose_account_concepts, propose_archetype,
                                                 propose_maturity_split)
    for fn in (propose_archetype, propose_account_concepts, propose_maturity_split):
        assert inspect.signature(fn).parameters["client"].default is None


# ----------------------------------------------------------------------------- upload a REAL TB end-to-end
def test_upload_tb_test_end_to_end_ties_to_published_fs():
    """Abhishek's real TB_Test.xlsx, THROUGH THE APP: parse → confirm columns → confirm sign → detect+
    confirm pre-close → map (preparer) → confirm → generate → faces tie to the published Draft FS. Roll-
    forward notes render 'not generated — no movement data', never faked. (AppTest can't drive
    st.file_uploader, so U1's load_grid + column proposal are pre-seeded into session_state — exactly the
    post-Parse state — then the confirm chain is driven; the file read itself is covered by test_tb_ingest.)"""
    from streamlit.testing.v1 import AppTest
    from ai_accountant.tb_ingest import detect_header_row, load_grid, propose_column_roles
    tb = ROOT / "tests" / "fixtures" / "tb_upload" / "tb_test.xlsx"
    if not tb.exists():
        print("[skip] tb_test fixture absent")
        return
    at = AppTest.from_file(str(APP), default_timeout=400).run()
    grid = load_grid(str(tb))
    hdr = detect_header_row(grid)
    at.session_state["mfs_grid"] = grid
    at.session_state["mfs_hdr"] = hdr
    at.session_state["mfs_colroles"] = propose_column_roles(grid[hdr], grid[hdr + 1:hdr + 6], client=None)
    at.session_state["mfs_mode"] = "upload"
    at.radio(key="mfs_source").set_value("Upload a TB (.xlsx / .csv)").run()
    at.button(key="mfs_confirm_cols").click().run()
          # U2 — accept heuristic column roles
    assert len(_ss(at, "mfs_parsed_tbrows")) == 79
    at.button(key="mfs_confirm_sign").click().run()
          # U3 — accept proposed credit-sign flips

    at.radio(key="mfs_pick").set_value("telecom_mobily").run()
    at.button(key="mfs_confirm_arch").click().run()
          # G1 — deterministic archetype, still confirmed
    assert _ss(at, "mfs_seed_id") == "telecom_mobily"
    # one-click "Confirm ALL AI proposals" (no per-line rerun) — the simple path
    assert "mfs_confirm_all" in [getattr(b, "key", None) for b in at.button]
    at.button(key="mfs_confirm_all").click().run()
           # G2 — preparer mappings confirmed in one click

    assert "mfs_confirm_close" in [getattr(b, "key", None) for b in at.button]   # C1 — pre-close DETECTED
    at.button(key="mfs_confirm_close").click().run()
    assert _ss(at, "mfs_close_done") is True
    at.selectbox(key="mfs_unit").set_value("SAR'000").run()


    m = _ss(at, "mfs_model")

    def L(cid):
        return next((l.current for st in m.statements.values() for l in st if l.concept_id == cid), None)

    assert round(L("tc_total_assets")) == 38515028                    # ties to the published Draft FS
    assert round(L("tc_net_profit")) == 3106848 and round(L("tc_total_ci")) == 3062385
    assert round(L("tc_retained")) == 11198161                        # opening 8,135,776 + result 3,062,385
    assert round(L("tc_total_assets") - L("tc_total_oe")) == 0        # SOFP balances
    assert not [f for f in m.findings if str(f[0]) == "balance check"]
    # S3b: PP&E builds as a two-layer static composition (NBV by class), with the MOVEMENT named absent in a
    # caveat (omit-don't-zero) — not silently faked, not a bare "not generated" status
    ppe = _note(m, "Property and equipment")
    assert ppe is not None and ppe.status in ("BUILT", "GROUPING_UNCONFIRMED")
    assert any("movement" in c.lower() and "not generated" in c.lower() for c in ppe.caveats)
    # B2a: the three asset pairs the TB can't split are FLAGGED 'split undetermined', not silently lumped
    su = {l.concept_id for st in m.statements.values() for l in st if l.split_undetermined}
    assert {"tc_contract_costs_nc", "tc_contract_assets_nc", "tc_fin_other_assets_nc"} <= su
    assert len([f for f in m.findings if str(f[0]).startswith("maturity:")]) == 3
    assert len(at.get("download_button")) == 2                        # export, after the unit
    # the comparative columns carry the TB's period all the way to the model — formatted as the FULL DATE
    # (header "31 Dec 2024 (SAR 000s)" → column "31st December 2024") — through the archetype-confirm restart
    # that must NOT wipe the upstream period (regression: it once did, and once showed "Current/Prior")
    assert m.period_current == "31st December 2024" and m.period_prior == "31st December 2023"
    assert not at.exception, at.exception


def test_agenda_gate_live_survey_path_runs_without_error():
    """Regression: the S3c agenda gate's LIVE-survey branch builds a ClientMappingStore with `MappingRecord`
    (fsgen_mfs.py) — a path the keyless tests never reach (survey_client is None → block skipped), so a missing
    import crashed only the real app. Force the live path with a fake client and assert it runs clean."""
    import fsgen_mfs
    from streamlit.testing.v1 import AppTest
    from ai_accountant.tb_ingest import detect_header_row, load_grid, propose_column_roles
    tb = ROOT / "tests" / "fixtures" / "tb_upload" / "tb_test.xlsx"
    if not tb.exists():
        print("[skip] tb_test fixture absent")
        return

    class _FakeLLM:                                          # NOT _DemoLLM — a one-off; tests only that the
        def complete_json(self, prompt, system=None, **k):  # MappingRecord/survey branch executes, not accuracy
            return {"buildable": ["tc_cash"], "not_buildable": [], "semantic_material": {"tc_cash": True}}

    saved = fsgen_mfs._real_client
    fsgen_mfs._real_client = lambda: _FakeLLM()              # force survey_client is not None
    try:
        at = AppTest.from_file(str(APP), default_timeout=400).run()
        grid = load_grid(str(tb))
        hdr = detect_header_row(grid)
        at.session_state["mfs_grid"] = grid
        at.session_state["mfs_hdr"] = hdr
        at.session_state["mfs_colroles"] = propose_column_roles(grid[hdr], grid[hdr + 1:hdr + 6], client=None)
        at.session_state["mfs_mode"] = "upload"
        at.radio(key="mfs_source").set_value("Upload a TB (.xlsx / .csv)").run()
        at.button(key="mfs_confirm_cols").click().run()
        at.button(key="mfs_confirm_sign").click().run()
        at.radio(key="mfs_pick").set_value("telecom_mobily").run()
        at.button(key="mfs_confirm_arch").click().run()
        at.button(key="mfs_confirm_all").click().run()
      # → reaches the agenda gate; the survey branch runs
        assert not at.exception, at.exception               # no NameError on MappingRecord
        assert _ss(at, "mfs_agenda_proposal") == {"buildable": ["tc_cash"], "not_buildable": [],
                                                  "semantic_material": {"tc_cash": True}}
        assert "mfs_confirm_agenda" in [getattr(b, "key", None) for b in at.button]
    finally:
        fsgen_mfs._real_client = saved                      # restore — the __main__ runner shares one process


# ----------------------------------------------------------------------------- static breakdown (Slice S1)
def _note(model, frag):
    return next((v for v in model.note_views if frag.lower() in v.note_ref.lower()), None)


def test_upload_static_breakdown_is_not_final_until_accepted_then_builds():
    """Mobily TB through the app: 'Financial and other assets' renders GROUPING_UNCONFIRMED (foots, but the
    grouping is unreviewed) until C2 accept, then BUILT and footing — while the PP&E roll-forward in the same
    FS stays 'not generated' (the no-GL boundary holds)."""
    from streamlit.testing.v1 import AppTest
    from ai_accountant.tb_ingest import detect_header_row, load_grid, propose_column_roles
    tb = ROOT / "tests" / "fixtures" / "tb_upload" / "tb_test.xlsx"
    if not tb.exists():
        print("[skip] tb_test fixture absent")
        return
    at = AppTest.from_file(str(APP), default_timeout=400).run()
    grid = load_grid(str(tb))
    hdr = detect_header_row(grid)
    at.session_state["mfs_grid"] = grid
    at.session_state["mfs_hdr"] = hdr
    at.session_state["mfs_colroles"] = propose_column_roles(grid[hdr], grid[hdr + 1:hdr + 6], client=None)
    at.session_state["mfs_mode"] = "upload"
    at.radio(key="mfs_source").set_value("Upload a TB (.xlsx / .csv)").run()
    at.button(key="mfs_confirm_cols").click().run()
    at.button(key="mfs_confirm_sign").click().run()
    at.radio(key="mfs_pick").set_value("telecom_mobily").run()
    at.button(key="mfs_confirm_arch").click().run()
    at.button(key="mfs_confirm_all").click().run()
    at.button(key="mfs_confirm_close").click().run()
    # C2 present; BEFORE accept the breakdown is not-final
    assert "mfs_accept_breakdowns" in [getattr(b, "key", None) for b in at.button]
    fa = _note(_ss(at, "mfs_model"), "Financial and other assets")
    assert fa is not None and fa.status == "GROUPING_UNCONFIRMED"
    at.button(key="mfs_accept_breakdowns").click().run()
        # explicit human accept
    m = _ss(at, "mfs_model")
    fa = _note(m, "Financial and other assets")
    assert fa.status == "BUILT"
    tot = fa.rows[-1]                                            # the grand total, shown tied to the face line
    assert tot.kind == "total" and round(tot.raw_sar) == 1001643 and "agrees to" in tot.label
    # S3a: the plain + contra static notes the TB supports all render BUILT, each tied to its face value
    nv = {v.note_ref: v for v in m.note_views}
    for cap, exp in (("Cash and cash equivalents", 1399542), ("Contract costs", 364218),
                     ("Due from related parties", 107332), ("Accounts receivable (net of allowance)", 3929559),
                     ("Inventories (net of provision)", 212992), ("Contract assets (net of allowance)", 1093454)):
        assert nv[cap].status == "BUILT", (cap, nv[cap].status)
        assert round(nv[cap].rows[-1].raw_sar) == exp           # contra notes net (gross + negative allowance)
    subs = {r.label: round(r.raw_sar) for r in fa.rows if r.kind == "subtotal"}   # two source tiers
    assert "Financial assets" in subs and "Other assets" in subs
    others = [round(r.raw_sar) for r in fa.rows if r.label == "Others"]           # the two "Others" stay SEPARATE
    assert len(others) == 2 and sum(others) == 257715           # 119,670 + 138,045, not fused into one line
    ppe = _note(m, "Property and equipment")                              # S3b: two-layer static, BUILT after accept
    assert ppe.status == "BUILT" and round(ppe.rows[-1].raw_sar) == 18851032   # NBV total foots to the face
    assert any("movement" in c.lower() and "not generated" in c.lower() for c in ppe.caveats)   # movement omitted, named
    at.selectbox(key="mfs_unit").set_value("SAR'000").run()
    assert len(at.get("download_button")) == 2 and not at.exception


def test_bank_static_breakdown_builds_cross_archetype():
    """The bank archetype (fixture path): 'Investments' breaks down into its measurement-category components —
    proves the static mechanism is archetype-agnostic (declared in the bank seed, built on the bank TB)."""
    at = _open_view("bank")
    _confirm_archetype(at, "ksa_bank")
    for s in [s for s in at.selectbox if str(getattr(s, "key", "")).startswith("mfs_map_")]:
        s.set_value(_CONFIRM)
    at.run()
    at.button(key="mfs_apply_map").click().run()
    assert "mfs_accept_breakdowns" in [getattr(b, "key", None) for b in at.button]
    at.button(key="mfs_accept_breakdowns").click().run()
    inv = _note(_ss(at, "mfs_model"), "Investments")
    assert inv is not None and inv.status == "BUILT"
    leaf_lines = [r for r in inv.rows if r.kind == "line"]      # one-tier (no source level): FVIS / FVOCI / amortised
    assert len(leaf_lines) == 3
    tot = inv.rows[-1]
    assert tot.kind == "total" and round(tot.raw_sar) == 22000 and "agrees to" in tot.label   # tied to the face


# ----------------------------------------------------------------------------- BS-split (Slice B2b) end-to-end GUI
def _contract_lumps(tb_path):
    """Compute the TB's LUMPED contract current/non-current totals via the engine's own resolve path (what the
    app does at G2), so the synthetic BS sheet can be built to reconcile to them."""
    from ai_accountant.master_fs import (ClientMappingStore, ProvenanceStore, apply_mapping_decisions,
                                         load_master_store)
    from ai_accountant.master_fs.orchestrator import _stored_from_mapping
    from ai_accountant.reporting.master_fs_export import build_master_fs_export
    from ai_accountant.tb_ingest import (apply_tb_sign, confirm_column_roles, confirm_tb_sign,
                                         detect_header_row, load_grid, parse_tb, propose_column_roles,
                                         propose_tb_sign, to_engine_inputs)
    from ai_accountant.tb_ingest.resolve import build_upload_proposals
    g = load_grid(str(tb_path))
    h = detect_header_row(g)
    schema = confirm_column_roles(propose_column_roles(g[h], g[h + 1:h + 6], client=None),
                                  {c.index: c.role for c in propose_column_roles(g[h], g[h + 1:h + 6], client=None)})
    items, rows = to_engine_inputs(parse_tb(g, schema, header_row=h))
    sp = propose_tb_sign(rows)
    rows = apply_tb_sign(rows, confirm_tb_sign(sp, {s.section: s.proposed for s in sp.sections}, approver="t", at="t"))
    store = load_master_store(seed_id="telecom_mobily")
    props = build_upload_proposals(rows, store, client=None)
    ms, au = ClientMappingStore(), ProvenanceStore()
    apply_mapping_decisions(props, store, client_id="uploaded-tb",
                            decisions={p.code: "confirm" for p in props}, mapping_store=ms, audit=au,
                            approver="t", at="t")
    m = build_master_fs_export(store, _stored_from_mapping(ms, "uploaded-tb", store.master_id, rows))
    out = {}
    for cid in ("tc_contract_costs_nc", "tc_contract_assets_nc", "tc_fin_other_assets_nc"):
        l = next((x for st in m.statements.values() for x in st if x.concept_id == cid), None)
        if l is not None:
            out[cid] = (l.current, l.prior)
    return out


def _multisheet_bytes(tb_path, bs_rows):
    """A 2-sheet workbook: 'TB Mapping' = the real tb_test content, 'Balance Sheet' = the synthetic split."""
    import io

    import openpyxl

    from ai_accountant.tb_ingest import load_grid
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    tb = wb.create_sheet("TB Mapping")
    for row in load_grid(str(tb_path)):
        tb.append(list(row))
    bs = wb.create_sheet("Balance Sheet")             # a 'balance sheet'-named sheet → best_bs_sheet finds it
    for row in bs_rows:
        bs.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_bs_split_end_to_end_through_the_gui_ties_the_face():
    """Slice B2b through the ACTUAL app: a multi-sheet upload (TB + balance-sheet face) drives U4 (pick + confirm
    the BS sheet), then the visible-reconcile gate; confirming APPLIES the split and the face ties. Drives the
    real `_g0_upload` second-sheet read (the file_uploader bytes path AppTest can't click is pre-seeded into
    session_state exactly as `Parse` would)."""
    import io

    from streamlit.testing.v1 import AppTest

    from ai_accountant.tb_ingest import detect_header_row, load_grid, propose_column_roles
    tb = ROOT / "tests" / "fixtures" / "tb_upload" / "tb_test.xlsx"
    if not tb.exists():
        print("[skip] tb_test fixture absent")
        return
    lumps = _contract_lumps(tb)
    assert "tc_contract_costs_nc" in lumps                       # the TB lumps contract costs (the B2a case)
    # build a BS sheet that reconciles to each lump in BOTH years (split the prior proportionally to current)
    bs_rows = [["Particulars", "2024", "2023"]]
    label = {"tc_contract_costs_nc": "Contract costs", "tc_contract_assets_nc": "Contract assets",
             "tc_fin_other_assets_nc": "Financial and other assets"}
    for cid, (lc, lp) in lumps.items():
        lp = lp or 0.0
        nc_cur = round(lc * 0.1, 2)
        nc_pri = round(lp * 0.1, 2)
        bs_rows.append([f"{label[cid]} - non-current", nc_cur, nc_pri])
        bs_rows.append([f"{label[cid]} - current", round(lc - nc_cur, 2), round(lp - nc_pri, 2)])
    raw = _multisheet_bytes(tb, bs_rows)

    at = AppTest.from_file(str(APP), default_timeout=600).run()
    grid = load_grid(io.BytesIO(raw), filename="wb.xlsx", sheet="TB Mapping")
    at.session_state["mfs_upload_bytes"] = raw
    at.session_state["mfs_upload_name"] = "wb.xlsx"
    at.session_state["mfs_sheets"] = ["TB Mapping", "Balance Sheet"]
    at.session_state["mfs_sheet"] = "TB Mapping"
    at.session_state["mfs_grid"] = grid
    at.session_state["mfs_hdr"] = detect_header_row(grid)
    at.session_state["mfs_colroles"] = propose_column_roles(grid[detect_header_row(grid)],
                                                            grid[detect_header_row(grid) + 1:detect_header_row(grid) + 6],
                                                            client=None)
    at.session_state["mfs_mode"] = "upload"
    at.radio(key="mfs_source").set_value("Upload a TB (.xlsx / .csv)").run()
    at.button(key="mfs_confirm_cols").click().run()
    at.button(key="mfs_confirm_sign").click().run()
    # U4: best_bs_sheet defaults to 'Balance Sheet' → confirm its columns (the second-sheet read)
    assert "mfs_confirm_bscols" in [getattr(b, "key", None) for b in at.button]
    at.button(key="mfs_confirm_bscols").click().run()
    # G1 + G2
    at.radio(key="mfs_pick").set_value("telecom_mobily").run()
    at.button(key="mfs_confirm_arch").click().run()
    at.button(key="mfs_confirm_all").click().run()
    # the visible-reconcile gate appears → confirm & apply
    assert "mfs_confirm_bssplit" in [getattr(b, "key", None) for b in at.button], \
        "the BS-split reconcile gate did not appear"
    at.button(key="mfs_confirm_bssplit").click().run()
    # drive the remaining gates (close + unit) so the statements build
    if "mfs_confirm_close" in [getattr(b, "key", None) for b in at.button]:
        at.button(key="mfs_confirm_close").click().run()
    at.selectbox(key="mfs_unit").set_value("SAR'000").run()
    assert not at.exception, at.exception

    m = _ss(at, "mfs_model")

    def cur(cid):
        return next((l.current for st in m.statements.values() for l in st if l.concept_id == cid), None)

    # the face now carries BOTH halves, summing to the lump; split_undetermined cleared for contract costs
    nc, c = cur("tc_contract_costs_nc"), cur("tc_contract_costs_c")
    assert nc is not None and c is not None and round(nc + c, 2) == round(lumps["tc_contract_costs_nc"][0], 2)
    su = {l.concept_id for st in m.statements.values() for l in st if l.split_undetermined}
    assert "tc_contract_costs_nc" not in su                       # applied → flag cleared
    assert "face_split" in _ss(at, "mfs_stored") and _ss(at, "mfs_stored").get("face_split_confirmed")


if __name__ == "__main__":
    failures = 0
    for name, fn in sorted(globals().items()):
        if name.startswith("test_") and callable(fn):
            try:
                fn()
                print(f"[pass] {name}")
            except AssertionError as exc:
                failures += 1
                print(f"[FAIL] {name}: {exc}")
            except Exception as exc:  # noqa: BLE001
                import traceback
                failures += 1
                print(f"[ERROR] {name}: {exc}")
                traceback.print_exc()
    print(f"\n{'ALL PASSED' if not failures else str(failures) + ' FAILED'}")
    sys.exit(1 if failures else 0)
