"""Slice SL-1c.2 — register enrichment, FINANCING-NET, reusing SL-1c.1's machinery UNCHANGED.

The headline is the REUSE proof on a REAL second register: Financing's DIFFERENT columns (gross/allowance/net +
segment/sector/stage) carry through the SAME `renderable_register` + emitters with ZERO render-code change. New
content is only the Financing specifics: NET is the tie column (gross/allowance carried as attributes), the
per-row gross+allowance==net cross-check (contra-as-COLUMN), and the honest prior-not-tied flag (the register has
no prior net — current-year ties, prior gross carried-not-reconciled, never faked).

    python tests/test_sub_ledger_register_financing.py
"""
from __future__ import annotations

import copy
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl                                                               # noqa: E402
from ai_accountant.tb_ingest.register import propose_register_concept, read_register   # noqa: E402
from ai_accountant.master_fs.seed import load_master_store                    # noqa: E402
from ai_accountant.reporting.master_fs_export import (build_master_fs_export,  # noqa: E402
                                                      export_master_fs_excel_bytes, export_master_fs_pdf_bytes)

ROOT = Path(__file__).resolve().parent.parent
REAL = ROOT / "Draft Data for working.xlsx"
FACE_CUR = 110862169                                                          # Σ Net 2025 == TB bs_loans (current)


def _reg():
    from ai_accountant.tb_ingest.grid import load_grid
    return read_register(load_grid(str(REAL), sheet="Financing -net"))


def _stored(store):
    tb = [{"account": "F1", "label": "Commercial Financing", "current": 70000000.0, "prior": 60000000.0,
           "section": "Assets", "maturity_hint": "", "levels": ["Assets", "", "Financing, net", "Commercial Financing", ""]},
          {"account": "F2", "label": "Consumer Financing", "current": 40862169.0, "prior": 36912496.0,
           "section": "Assets", "maturity_hint": "", "levels": ["Assets", "", "Financing, net", "Consumer Financing", ""]}]
    maps = [{"account": r["account"], "concept_id": "bs_loans", "client_label": r["label"], "provenance": "preparer"}
            for r in tb]
    return {"client": "draft", "master_id": store.master_id, "tb": tb, "mappings": maps, "extensions": []}


def _build(store, reg):
    return build_master_fs_export(store, _stored(store), gl={"bs_loans": reg}, notes_attempted=True)


# ============================================================ read: NET is the tie column, no prior, segments
def test_financing_read_picks_net_no_prior_segment_sections():
    if not REAL.exists():
        print("[skip]"); return
    reg = _reg()
    assert "net" in reg["amount_current_header"].lower() and "2025" in reg["amount_current_header"]
    assert reg["amount_prior_header"] is None                                # NO prior net → no prior tie column
    assert any("Gross" in c for c in reg["columns"]) and any("Allowance" in c for c in reg["columns"])
    assert len([s for s in reg["sections"] if s.startswith(("A)", "B)", "C)", "D)"))]) >= 4   # by SEGMENT
    assert round(sum(r["current"] for r in reg["rows"])) == FACE_CUR


# ============================================================ end-to-end: current tie BUILT, prior carried-not-tied
def test_financing_current_year_ties_prior_carried_not_tied():
    if not REAL.exists():
        print("[skip]"); return
    store = load_master_store(seed_id="ksa_bank")
    r = _build(store, _reg()).note_results["bs_loans"]["register"]
    assert r["status"] == "BUILT" and abs(r["total"] - r["face_value"]) < 0.01 and r["has_prior"] is False
    assert any("no prior net" in f[2] for f in r["findings"])                # the prior is honestly flagged not-tied
    assert not any("BLOCKS" in f[2] for f in r["findings"])                  # …but the note is NOT blocked (current ties)


# ============================================================ THE REUSE PROOF — different columns, same render
def test_reuse_render_carries_financing_columns_pdf_excel():
    if not REAL.exists():
        print("[skip]"); return
    store = load_master_store(seed_id="ksa_bank")
    m = _build(store, _reg())
    nv = next(v for v in m.note_views if v.columns)
    assert any("Segment" in c for c in nv.columns) and any("Gross" in c for c in nv.columns)   # Financing's OWN columns
    assert len(export_master_fs_pdf_bytes(m)) > 5000
    wb = openpyxl.load_workbook(io.BytesIO(export_master_fs_excel_bytes(m)))
    vals = [str(c.value) for s in wb.sheetnames for row in wb[s].iter_rows() for c in row if c.value]
    assert any("Murabaha" == v for v in vals) and any("Stage 1" == v for v in vals)            # attribute facts VERBATIM
    assert any("Segment" in v for v in vals) and any("agrees to" in v for v in vals)


# ============================================================ tie + per-row cross-check (numbers never on trust)
def test_broken_sum_blocks():
    if not REAL.exists():
        print("[skip]"); return
    store = load_master_store(seed_id="ksa_bank")
    reg = _reg(); reg["rows"] = reg["rows"][:-1]
    assert _build(store, reg).note_results["bs_loans"]["register"]["status"] == "BLOCKED"


def test_per_row_gross_allowance_net_mismatch_flags_even_when_the_sum_ties():
    if not REAL.exists():
        print("[skip]"); return
    store = load_master_store(seed_id="ksa_bank")
    reg = copy.deepcopy(_reg())
    g_i = next(i for i, c in enumerate(reg["columns"]) if "Gross" in c and "2025" in c)
    reg["rows"][0]["cells"][g_i] = str(_pnum(reg["rows"][0]["cells"][g_i]) + 5000)   # break gross only → net SUM untouched
    r = _build(store, reg).note_results["bs_loans"]["register"]
    assert abs(r["total"] - r["face_value"]) < 0.01                          # the net SUM still ties …
    assert r["status"] == "BLOCKED" and any("gross + allowance" in f[2] for f in r["findings"])  # …per-row catches it


def _pnum(v):
    from ai_accountant.tb_ingest.parse import _num
    return _num(v)


# ============================================================ coarse-vs-enriched + association + regression
def test_coarse_vs_enriched_and_association_and_investment_still_ties():
    store = load_master_store(seed_id="ksa_bank")
    # association: the Financing sheet resolves at Layer 0 (deterministic, no model)
    out = propose_register_concept("Financing -net", ["Loan/Facility ID", "Segment", "Financing, net 2025"],
                                   [("bs_loans", "Loans / financing, net")], client=None)
    assert out["concept_id"] == "bs_loans" and out["layer"] == 0
    # absent → the coarse bs_loans static note (no register) — but bs_loans declares contra_ecl too; with no gl it
    # builds the static breakdown floor. Present → enriched register.
    if REAL.exists():
        assert "register" in _build(store, _reg()).note_results["bs_loans"]
    # REGRESSION: SL-1c.1's Investment register still ties (the bs_loans addition didn't disturb it)
    if REAL.exists():
        inv = read_register(__import__("ai_accountant.tb_ingest.grid", fromlist=["load_grid"]).load_grid(
            str(REAL), sheet="Investment "))
        assert round(sum(r["current"] for r in inv["rows"])) == 38967880


if __name__ == "__main__":
    failures = 0
    for name, fn in sorted(globals().items()):
        if name.startswith("test_") and callable(fn):
            try:
                fn()
                print(f"[pass] {name}")
            except AssertionError as exc:
                failures += 1
                print(f"[FAIL] {name}: {str(exc).encode('ascii', 'replace').decode()}")
            except Exception as exc:  # noqa: BLE001
                import traceback
                failures += 1
                print(f"[ERROR] {name}: {exc}")
                traceback.print_exc()
    print(f"\n{'ALL PASSED' if not failures else str(failures) + ' FAILED'}")
    sys.exit(1 if failures else 0)
