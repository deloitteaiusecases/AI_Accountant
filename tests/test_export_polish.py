"""Slice E1: export polish (PDF + Excel) — the proof the user asked for, frozen as guards.

  * DEDUPE — a dual-declared concept (roll_forward + static_breakdown) renders ONCE, and the surviving
    view is the BUILT one: telecom (TB-only) keeps the STATIC composition; the bank (GL) keeps the
    ROLL_FORWARD movement (GUARD 1 — dedupe must respect the S3b GL-presence pick, never drop a real note).
  * YEAR LABELS — the comparative columns read the TB's real period header (verbatim), not "Current/Prior".
  * BRACKET NEGATIVES — a negative reads as (1,234) in the PDF; in Excel the cell STAYS a real negative
    number (summable) and only its number_format paints the brackets (GUARD 2 — never a "(200)" string).
  * ORDER — the AI's working (mappings / findings) sits AFTER the note pages in the PDF.

    python tests/test_export_polish.py
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import ai_accountant.master_fs                                                          # noqa: E402,F401  (package-first import)
from ai_accountant.reporting.master_fs_export import (_acct, build_master_fs_export,    # noqa: E402
                                                      export_master_fs_excel_bytes,
                                                      export_master_fs_pdf_bytes)
from tests.test_maturity_split import _mobily_stored                                    # noqa: E402

_CUR, _PRI = "31 Dec 2024 (SAR 000s)", "31 Dec 2023 (SAR 000s)"


def _pdf_pages(model):
    try:
        from pypdf import PdfReader
    except ImportError:
        from PyPDF2 import PdfReader
    rd = PdfReader(io.BytesIO(export_master_fs_pdf_bytes(model)))
    return [pg.extract_text() or "" for pg in rd.pages]


def _mobily_model():
    store, stored = _mobily_stored()
    return build_master_fs_export(store, stored, notes_attempted=True,
                                  period_current=_CUR, period_prior=_PRI)


# ----------------------------------------------------------------------------- the pure formatter
def test_acct_brackets_negatives_only():
    assert _acct(-564681, 2) == "(564,681.00)"
    assert _acct(1399542, 2) == "1,399,542.00"
    assert _acct(-200, 0) == "(200)"
    assert _acct(0, 0) == "0"                                  # zero is not negative — no brackets
    assert _acct(None) == "—"                                 # withheld / absent


# ----------------------------------------------------------------------------- DEDUPE (GUARD 1)
def test_dual_declared_concept_renders_once_telecom_keeps_static():
    model = _mobily_model()
    refs = [v.note_ref for v in model.note_views]
    assert len(refs) == len(set(refs)), f"duplicate note views: {refs}"
    ppe = [v for v in model.note_views if v.note_ref == "Property and equipment"]
    assert len(ppe) == 1                                       # ONCE — not twice (roll_forward + static decls)
    # TB-only (no GL) → the surviving view is the STATIC composition (no movement "Opening" row)
    assert not any(r.label.lower() == "opening" for r in ppe[0].rows)
    assert ppe[0].rows and ppe[0].status in ("BUILT", "GROUPING_UNCONFIRMED")


def test_dual_declared_concept_renders_once_bank_keeps_rollforward():
    from tests.test_bank_notes_demo import _setup
    from ai_accountant.master_fs import generate_master_fs
    bstore, bstored, payloads = _setup()
    res = generate_master_fs(bstored, seed_id="ksa_bank", client="x", strategy="replay", bank="b", gl=payloads)
    refs = [v.note_ref for v in res.model.note_views]
    assert len(refs) == len(set(refs)), f"duplicate note views: {refs}"
    ppe = [v for v in res.model.note_views if v.note_ref.lower().startswith("property")]
    assert len(ppe) == 1 and ppe[0].status == "BUILT"
    # the bank has a GL → the surviving view is the ROLL_FORWARD movement (it carries an Opening row), NOT a
    # deduped-away static stub — the dedupe respected the GL-presence pick.
    assert any(r.label.lower() == "opening" for r in ppe[0].rows)


# ----------------------------------------------------------------------------- YEAR LABELS + BRACKETS in the PDF
def test_pdf_note_columns_read_the_real_years_and_negatives_are_bracketed():
    pages = _pdf_pages(_mobily_model())
    ppe = next(p for p in pages if "Property and equipment" in p and "Acc Dep" in p)
    assert _CUR in ppe and _PRI in ppe                         # the TB's own period header, verbatim — not "Current"
    assert "Current (" not in ppe and "Prior (" not in ppe     # the generic fallback is NOT used when a label exists
    assert "(564,681.00)" in ppe                               # the accumulated-depreciation contra, in brackets
    assert "-564,681" not in ppe                               # never a bare minus


def test_pdf_puts_mappings_and_findings_after_the_notes():
    pages = _pdf_pages(_mobily_model())

    def first(sub):
        return next((i for i, p in enumerate(pages) if sub in p), None)
    first_note = next(i for i, p in enumerate(pages) if "Property and equipment" in p and "Acc Dep" in p)
    assert first("Mappings applied") > first_note              # the AI's working comes AFTER the statements + notes
    assert first("Open findings") > first_note


def test_not_generated_note_shows_the_face_line_not_a_placeholder():
    """A note whose breakdown can't be built (here lease liabilities — roll-forward, no GL) now shows the
    figure(s) CARRIED TO THE FACE statement (the word-for-word FS line, under the year columns) instead of the
    old 'face line is unaffected; this note was not generated' placeholder. Still honest: status 'not generated'."""
    model = _mobily_model()
    ng = [v for v in model.note_views if v.status == "not generated" and v.rows]
    assert ng, "expected a not-generated note that surfaces its face line"
    face_vals = {round(l.current, 2) for st in model.statements.values() for l in st if l.current is not None}
    for v in ng:
        assert all(round(r.value, 2) in face_vals for r in v.rows)   # the face figure(s), never a fabricated breakdown
        assert any(r.prior_value is not None for r in v.rows)        # both years carried
    pages = _pdf_pages(model)
    ngpage = next(p for p in pages if "STATUS: not generated" in p and "Lease liabilities" in p)
    assert "face line is unaffected" not in ngpage                  # the old empty placeholder is gone
    assert "2,061,787.00" in ngpage                                 # the actual face figure is shown instead


def test_note_total_tie_line_wraps_intact_not_overprinted():
    """Regression: the long total tie-line ("Total — agrees to <concept> on the Statement of Financial
    Position") is a Paragraph in the note table, so it WRAPS inside the label column rather than overprinting
    the amount columns (the in-note PDF overlap). The full phrase + both amounts must survive intact."""
    pages = _pdf_pages(_mobily_model())
    ppe = next(p for p in pages if "Property and equipment" in p and "Acc Dep" in p)
    assert "agrees to Property and equipment on the" in ppe     # the tie-line text is intact (not garbled/clipped)
    assert "Statement of Financial Position" in ppe
    assert "18,851,032.00" in ppe and "19,011,971.00" in ppe    # its current + prior amounts both present


def test_findings_long_account_id_wraps_intact_beside_its_label():
    """Regression: a long, space-less finding key (e.g. maturity:tc_contract_costs_nc) is rendered in a
    Paragraph so it WRAPS inside the Account column — it must appear intact, never garbled by overprinting
    the Client-label column (the PDF table-overlap bug)."""
    pages = _pdf_pages(_mobily_model())
    fp = next(p for p in pages if "Open findings" in p)
    assert "maturity:tc_contract_costs_nc" in fp                # the full key survives intact (not overprinted)
    assert "Contract costs" in fp                               # its label is present and separate


# ----------------------------------------------------------------------------- BRACKETS in Excel stay NUMERIC (GUARD 2)
def test_excel_negatives_stay_numeric_with_a_bracket_number_format():
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(export_master_fs_excel_bytes(_mobily_model())))
    ws = next(wb[n] for n in wb.sheetnames if "Property" in n)
    negs = [c for row in ws.iter_rows() for c in row if isinstance(c.value, (int, float)) and c.value < 0]
    assert negs, "expected a negative (accumulated depreciation) cell in the PP&E note"
    for c in negs:                                             # a REAL number (summable), brackets only DISPLAY
        assert isinstance(c.value, (int, float)) and not isinstance(c.value, str)
        assert c.number_format == "#,##0.00;(#,##0.00)"
    # and the column header is the real year, used verbatim (no doubled unit)
    headers = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str) and _CUR in c.value]
    assert headers and all(h == _CUR for h in headers)


if __name__ == "__main__":
    failures = 0
    for name, fn in sorted(globals().items()):
        if name.startswith("test_") and callable(fn):
            try:
                fn()
                print(f"[pass] {name}")
            except AssertionError as exc:
                failures += 1
                print(f"[FAIL] {name}: {exc}")
            except Exception as exc:  # noqa: BLE001
                import traceback
                failures += 1
                print(f"[ERROR] {name}: {exc}")
                traceback.print_exc()
    print(f"\n{'ALL PASSED' if not failures else str(failures) + ' FAILED'}")
    sys.exit(1 if failures else 0)
