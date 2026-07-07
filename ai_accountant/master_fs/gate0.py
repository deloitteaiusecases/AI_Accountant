"""GATE 0 — prove a seed's JSON faithfully represents its human-approved xlsx (and the CSV the JSON).

This is BANK-APPROVED-XLSX RECONCILIATION, not the seed-driven runtime path: it only runs for a seed that
HAS an approved xlsx (the bank does), reads it for VALIDATION ONLY (a colour-coded human reference, never
runtime data), and necessarily references that xlsx's alias columns (`aljazira` / `saab`). A new archetype
seed without an xlsx never invokes it. The headline check is the per-concept ALIAS PAIRING: for each
`both_naming_differs` concept the two labels must be filed under the RIGHT bank — "both strings present"
is not "correctly paired" (a swap would render every client FS with the wrong bank's wording).
"""
from __future__ import annotations

import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook

# The approved-xlsx TEMPLATE marks each statement sheet with this prefix (a format marker, NOT an
# archetype fact). Statement sheets are mapped, in workbook order, to the seed's DECLARED statement keys.
_SHEET_PREFIX = "Master — "


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def _norm_section(s) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(s or "").strip().lower()).strip("_")


def _label(s) -> str:
    """A bank's label, with the xlsx 'not present' placeholders (— / –) normalised to empty."""
    t = _norm(s)
    return "" if t in ("—", "–", "-", "") else t


def _norm_presence(s) -> str:
    """Map BOTH the JSON codes and the human xlsx text to one canonical presence code, so a faithful
    JSON (codes) validates against the human reference (prose) without false mismatches."""
    t = str(s or "").strip().lower()
    if "naming differs" in t or t == "both_naming_differs":
        return "both_naming_differs"
    if t.startswith("aljazira") or t == "aljazira_only":
        return "aljazira_only"
    if t.startswith("saab") or t == "saab_only":
        return "saab_only"
    if t == "both":
        return "both"
    return re.sub(r"[^a-z0-9]+", "_", t).strip("_")


def extract_xlsx(xlsx_path, statement_keys) -> dict:
    """{statement_key: [ {section, group, canonical, aljazira, saab, presence, order} ... ]}. Statement
    SHEETS (template-prefixed) are zipped, in workbook order, to the seed's declared `statement_keys`."""
    wb = load_workbook(xlsx_path, data_only=True)
    sheets = [s for s in wb.sheetnames if s.startswith(_SHEET_PREFIX)]
    out: dict = {}
    for sheet, statement in zip(sheets, statement_keys):
        ws = wb[sheet]
        rows, section, order = [], None, 0
        for r in range(5, ws.max_row + 1):
            c = [ws.cell(r, i).value for i in range(1, 7)]
            canonical = _norm(c[2])
            if not canonical:                              # section/blank row
                if _norm(c[0]):
                    section = _norm(c[0])
                continue
            order += 10
            rows.append({"section": section, "group": _norm(c[1]) or None, "canonical": canonical,
                         "aljazira": _norm(c[3]), "saab": _norm(c[4]), "presence": _norm(c[5]),
                         "order": order})
        out[statement] = rows
    return out


def validate_seed(json_path, xlsx_path) -> list:
    """Returns a list of human-readable mismatches; empty list == the seed faithfully matches the xlsx."""
    data = json.loads(Path(json_path).read_text(encoding="utf-8"))
    xlsx = extract_xlsx(xlsx_path, list(data["statements"].keys()))
    problems: list = []

    for statement, jrows in data["statements"].items():
        xrows = xlsx.get(statement, [])
        # a DRAFT-ONLY concept (e.g. an estimated balancing residual a specific client TB carries) is by
        # definition NOT part of the approved PUBLISHED FS — exclude it from the published-FS fidelity check,
        # so it isn't flagged as an "invention". The CSV flat-copy still carries it (JSON↔CSV stays in sync).
        jrows = [r for r in jrows if _norm(r.get("presence")) != "draft_only"]
        jx = {_norm(r["canonical_concept"]): r for r in jrows}
        xx = {r["canonical"]: r for r in xrows}

        for missing in sorted(set(jx) - set(xx)):
            problems.append(f"[{statement}] JSON concept not in approved xlsx (invention?): {missing!r}")
        for dropped in sorted(set(xx) - set(jx)):
            problems.append(f"[{statement}] approved xlsx line missing from JSON (dropped): {dropped!r}")

        if [r["canonical_concept"] for r in jrows] != [r["canonical"] for r in xrows]:
            problems.append(f"[{statement}] concept ORDER differs between JSON and the approved xlsx")

        for canonical, jr in jx.items():
            xr = xx.get(canonical)
            if xr is None:
                continue
            if _norm_section(jr["l0_section"]) != _norm_section(xr["section"]):
                problems.append(f"[{statement}] {canonical!r}: section {jr['l0_section']!r} != "
                                f"xlsx {xr['section']!r}")
            if _norm_presence(jr.get("presence")) != _norm_presence(xr["presence"]):
                problems.append(f"[{statement}] {canonical!r}: presence {jr.get('presence')!r} != "
                                f"xlsx {xr['presence']!r}")
            # THE ALIAS PAIRING — each bank's label filed under the RIGHT bank (— = not present)
            aliases = jr.get("label_aliases", {})
            if _label(aliases.get("aljazira")) != _label(xr["aljazira"]):
                problems.append(f"[{statement}] {canonical!r}: aljazira label "
                                f"{aliases.get('aljazira')!r} != xlsx {xr['aljazira']!r} (PAIRING)")
            if _label(aliases.get("saab")) != _label(xr["saab"]):
                problems.append(f"[{statement}] {canonical!r}: saab label {aliases.get('saab')!r} != "
                                f"xlsx {xr['saab']!r} (PAIRING)")
    return problems


def validate_csv(json_path, csv_path) -> list:
    """The flat CSV must equal the JSON (so a diff-friendly copy can't drift)."""
    data = json.loads(Path(json_path).read_text(encoding="utf-8"))
    j = {r["concept_id"]: r for rows in data["statements"].values() for r in rows}
    problems: list = []
    seen = set()
    with open(csv_path, encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            cid = row["concept_id"]
            seen.add(cid)
            jr = j.get(cid)
            if jr is None:
                problems.append(f"CSV concept_id {cid!r} not in JSON")
                continue
            if _norm(row["canonical_concept"]) != _norm(jr["canonical_concept"]):
                problems.append(f"{cid}: CSV canonical != JSON")
            if _norm(row["aljazira_label"]) != _norm(jr["label_aliases"].get("aljazira")):
                problems.append(f"{cid}: CSV aljazira_label != JSON alias")
            if _norm(row["saab_label"]) != _norm(jr["label_aliases"].get("saab")):
                problems.append(f"{cid}: CSV saab_label != JSON alias")
    for cid in sorted(set(j) - seen):
        problems.append(f"JSON concept_id {cid!r} missing from CSV")
    return problems
