"""Normalise an uploaded TB (xlsx OR csv) to ONE grid (list of rows) before any role detection.

A CSV is just rows with no sheets/formatting, so it normalises to the same grid an xlsx first-sheet
yields — the column-role proposer reasons on headers + sample rows identically. Single workbook /
single sheet only (the GL/multi-sheet slice owns the rest)."""
from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any


def _is_csv(file: Any, filename: "str | None") -> bool:
    name = filename or getattr(file, "name", "") or (str(file) if isinstance(file, (str, Path)) else "")
    return str(name).lower().endswith(".csv")


def sheet_names(file: Any, *, filename: "str | None" = None) -> list:
    """The workbook's sheet names ([] for a csv). Used to let the user pick WHICH sheet is the TB when a
    workbook has several (TB / BS / notes …) — combining multiple sheets is the parked slice."""
    if _is_csv(file, filename):
        return []
    import openpyxl
    if hasattr(file, "seek"):
        file.seek(0)
    wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def best_tb_sheet(names: list) -> "str | None":
    """Best-guess which sheet is the trial balance, by name (a 'TB' / 'trial' / 'mapping' sheet), else the
    first. The user confirms the pick in the UI — this is only the default."""
    if not names:
        return None
    for n in names:
        low = str(n).lower()
        if "tb" in low or "trial" in low or "mapping" in low:
            return n
    return names[0]


# tokens that SUGGEST a balance-sheet / statement-of-financial-position face sheet — SCORED, not a single
# literal (no 'BS' hardcode); the user confirms the pick, this is only the default. The current/non-current
# split a TB lumps lives on this face sheet (Slice B2b).
_BS_TOKENS = (("statement of financial position", 3), ("financial position", 3), ("sofp", 3),
              ("balance sheet", 3), ("balancesheet", 3), (" bs ", 2), ("bs", 1), ("b/s", 2))


def best_bs_sheet(names: list, *, exclude: "str | None" = None) -> "str | None":
    """Best-guess which sheet is the BALANCE-SHEET face (carrying the current/non-current split), by a SCORED
    name match — never a single hardcoded 'BS' literal. `exclude` (the already-picked TB sheet) is skipped so
    a 'TB Mapping' sheet isn't also read as the BS. No confident match → None (→ the user picks, or no split)."""
    best, best_score = None, 0
    for n in names:
        if exclude is not None and n == exclude:
            continue
        low = f" {str(n).lower()} "
        score = max((w for tok, w in _BS_TOKENS if tok in low), default=0)
        if score > best_score:
            best, best_score = n, score
    return best


def load_grid(file: Any, *, filename: "str | None" = None, sheet: "str | None" = None) -> list:
    """Return the TB as a list-of-rows grid. `file` is a path, bytes, or a file-like (Streamlit upload).
    xlsx → the named `sheet` (or the first if None); csv → decoded rows. Cells are kept as-is (str/number);
    `detect_header_row` finds the header downstream."""
    if _is_csv(file, filename):
        if isinstance(file, (str, Path)):
            text = Path(file).read_text(encoding="utf-8-sig")
        else:
            raw = file.read() if hasattr(file, "read") else file
            text = raw.decode("utf-8-sig") if isinstance(raw, (bytes, bytearray)) else str(raw)
        return [list(r) for r in csv.reader(io.StringIO(text))]

    import openpyxl
    if hasattr(file, "seek"):
        file.seek(0)
    wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
    ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.worksheets[0]
    grid = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return grid
