"""Excel export — status on the SAME sheet as the numbers, frozen so it is visible without scrolling.

Not on a tab nobody clicks: the status block sits in the top rows and is frozen, and the total row
carries the caveat inline. A reader who opens the sheet and never scrolls still sees PARTIAL /
MAGNITUDE UNVERIFIED next to the figures.
"""
from __future__ import annotations

import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ai_accountant.reporting.render_model import renderable

_RED = PatternFill("solid", fgColor="B00020")
_WHITE_BOLD = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)


def _sheet_name(note_ref: str) -> str:
    name = re.sub(r"[:\\/?*\[\]]", " ", note_ref).strip()
    return name[:31] or "Note"


_ACCT_FMT = "#,##0.00;(#,##0.00)"                  # negatives in brackets — DISPLAY only; the cell stays a number


def write_note_sheet(wb: Workbook, rn, status_fill=_RED, subtle=False,
                     period_current=None, period_prior=None) -> None:
    """Render ONE renderable note as a sheet (status block frozen on top, caveats inline). Shared by
    the notes-alone export and the statement-first export so both carry identical honesty. The status WORD
    is always written, never colour-only. `subtle=True` (TEMPORARY demo styling): render the status as a
    COLOURED FOOTNOTE word (no heavy fill) — `status_fill` is then a hex colour string for the text; the
    default `subtle=False` keeps the bold red fill (the note-pipeline look, unchanged).
    `period_current`/`period_prior` label the comparative columns with the TB's real years (else Current/Prior)."""
    ws = wb.create_sheet(_sheet_name(rn.note_ref))
    row = 1
    _final = rn.status in ("BUILT", "RECONCILED", "not generated")
    txt = f"STATUS: {rn.status}" + ("" if _final else "  (NOT FINAL)")
    c = ws.cell(row, 1, txt)
    if subtle:
        c.font = Font(bold=True, color=status_fill if isinstance(status_fill, str) else "B00020")
    else:
        c.font = _WHITE_BOLD
        c.fill = status_fill
    row += 1
    ws.cell(row, 1, rn.status_line).font = _BOLD
    row += 1
    for cav in rn.caveats:                         # caveats in the always-visible top block
        cell = ws.cell(row, 1, f"!! {cav}")
        if subtle:
            cell.font = Font(italic=True, size=9, color=status_fill if isinstance(status_fill, str) else "B00020")
        else:
            cell.font = _WHITE_BOLD
            cell.fill = _RED
        row += 1
    row += 1
    if getattr(rn, "columns", None):                 # SL-1c.1 — REGISTER TABLE: N attribute columns + amount(s)
        _write_register_sheet(ws, rn, row)
        return
    has_prior = any(r.prior_value is not None for r in rn.rows)   # a comparative column (Slice S3b)
    ws.cell(row, 1, "Line").font = _BOLD
    # a real period label (the TB's own header) is used verbatim; "(unit)" only dresses the generic fallback
    if has_prior:
        ws.cell(row, 2, period_current or f"Current ({rn.unit_label})").font = _BOLD
        ws.cell(row, 3, period_prior or f"Prior ({rn.unit_label})").font = _BOLD
    else:
        ws.cell(row, 2, f"Amount ({rn.unit_label})").font = _BOLD
    header_row = row
    row += 1
    ws.freeze_panes = ws.cell(header_row + 1, 1)   # status block + header stay on screen

    for r in rn.rows:
        label = ("    " * r.indent) + r.label
        if r.flag and r.kind != "total":
            label += f"   [{r.flag.upper()}]"
        cell = ws.cell(row, 1, label)
        if r.kind == "total":
            mag = "  [MAGNITUDE UNVERIFIED]" if r.flag == "magnitude" else ""
            tag = "  — NOT FINAL" if rn.status != "BUILT" else ""
            cell.value = f"{r.label}{mag}{tag}"
            cell.font = _BOLD if rn.status == "BUILT" else _WHITE_BOLD
            if rn.status != "BUILT":
                cell.fill = _RED
        elif r.kind in ("section", "subtotal", "movement"):
            cell.font = _BOLD
        if r.kind != "movement":
            vcell = ws.cell(row, 2, r.value if r.value is not None else "WITHHELD")
            vcell.alignment = Alignment(horizontal="right")
            vcell.number_format = _ACCT_FMT            # bracket-negatives; numeric value untouched (GUARD 2)
            if has_prior:                              # withhold the prior column on the SAME R11 rule
                pv = r.prior_value if (rn.unit_confirmed and r.prior_value is not None) else (
                    "WITHHELD" if not rn.unit_confirmed else None)
                pc = ws.cell(row, 3, pv)
                pc.alignment = Alignment(horizontal="right")
                pc.number_format = _ACCT_FMT
        row += 1
    ws.column_dimensions["A"].width = 62
    ws.column_dimensions["B"].width = 30
    if has_prior:
        ws.column_dimensions["C"].width = 30


def _write_register_sheet(ws, rn, row) -> None:
    """SL-1c.1 — a register-enriched note as an N-COLUMN sheet: the attribute columns (`rn.columns`, VERBATIM
    `RenderRow.cells`) + the amount column(s). Column-LIST driven; numbers stay numeric (the bracket format is
    DISPLAY only). Section/sub-total/total rows bold the label; cells carry the source string untouched."""
    cols = list(rn.columns)
    has_prior = any(r.prior_value is not None for r in rn.rows)
    n = len(cols)
    for ci, h in enumerate(cols, start=1):
        ws.cell(row, ci, str(h).replace("\n", " ")).font = _BOLD
    ws.cell(row, n + 1, f"Current ({rn.unit_label})").font = _BOLD
    if has_prior:
        ws.cell(row, n + 2, f"Prior ({rn.unit_label})").font = _BOLD
    header_row = row
    ws.freeze_panes = ws.cell(header_row + 1, 1)
    row += 1
    for r in rn.rows:
        bold = r.kind in ("section", "subtotal", "total")
        if r.kind in ("section", "subtotal", "total") or not r.cells:   # a label row (no per-column cells)
            lc = ws.cell(row, 1, r.label)
            lc.font = _BOLD
        else:
            for ci, v in enumerate(r.cells[:n], start=1):
                ws.cell(row, ci, v)                       # VERBATIM attribute value — never reformatted
        if r.value is not None:
            vc = ws.cell(row, n + 1, r.value)
            vc.alignment = Alignment(horizontal="right"); vc.number_format = _ACCT_FMT
            if bold:
                vc.font = _BOLD
            if has_prior and r.prior_value is not None:
                pc = ws.cell(row, n + 2, r.prior_value)
                pc.alignment = Alignment(horizontal="right"); pc.number_format = _ACCT_FMT
                if bold:
                    pc.font = _BOLD
        row += 1
    for ci in range(1, n + 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = 20
    ws.column_dimensions[ws.cell(1, n + 1).column_letter].width = 18


def export_excel(notes: list, path: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)
    for note in notes:
        write_note_sheet(wb, renderable(note))
    wb.save(path)
    return path


def _export_excel_legacy(notes: list, path: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)
    for note in notes:
        rn = renderable(note)
        ws = wb.create_sheet(_sheet_name(rn.note_ref))
        row = 1
        c = ws.cell(row, 1, f"STATUS: {rn.status}" + ("  (NOT FINAL)" if rn.status != "BUILT" else ""))
        c.font = _WHITE_BOLD
        c.fill = _RED
        row += 1
        ws.cell(row, 1, rn.status_line).font = _BOLD
        row += 1
        for cav in rn.caveats:                         # caveats in the always-visible top block
            cell = ws.cell(row, 1, f"!! {cav}")
            cell.font = _WHITE_BOLD
            cell.fill = _RED
            row += 1
        row += 1
        ws.cell(row, 1, "Line").font = _BOLD
        ws.cell(row, 2, f"Amount ({rn.unit_label})").font = _BOLD
        header_row = row
        row += 1
        ws.freeze_panes = ws.cell(header_row + 1, 1)   # status block + header stay on screen

        for r in rn.rows:
            label = ("    " * r.indent) + r.label
            if r.flag and r.kind != "total":
                label += f"   [{r.flag.upper()}]"
            cell = ws.cell(row, 1, label)
            if r.kind == "total":
                mag = "  [MAGNITUDE UNVERIFIED]" if r.flag == "magnitude" else ""
                tag = "  — NOT FINAL" if rn.status != "BUILT" else ""
                cell.value = f"{r.label}{mag}{tag}"
                cell.font = _BOLD if rn.status == "BUILT" else _WHITE_BOLD
                if rn.status != "BUILT":
                    cell.fill = _RED
            elif r.kind in ("section", "subtotal", "movement"):
                cell.font = _BOLD
            if r.kind != "movement":
                vcell = ws.cell(row, 2, r.value if r.value is not None else "WITHHELD")
                vcell.alignment = Alignment(horizontal="right")
            row += 1
        ws.column_dimensions["A"].width = 62
        ws.column_dimensions["B"].width = 30
    wb.save(path)
    return path
