"""Master-FS export — detailed PDF/Excel of the phase-1 master-FS output.

Reuses the honesty machinery (provenance rendering, the findings page, the conditional NOT-FINAL stamp,
"never claim more certainty than the engine assigned") but points it at the master-FS content: the three
face statements from the master (populated lines, master order, client labels, comparative columns), a
mappings-applied/provenance view, and a findings page. Note-by-note breakdowns are a LATER phase — the
document shows the faces only and SAYS so; it never renders a placeholder note. The NOT-FINAL stamp is
conditional on real state (any AI-assumed line / unmapped item) and rendered as a small plain footnote,
not a banner. No illustrative banner — the figures are clearly illustrative round numbers.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from functools import partial

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from ai_accountant.master_fs.model import ClientMappingStore, MappingRecord, PREPARER
from ai_accountant.master_fs.render import render_comparative

_PROV_LABEL = {"preparer": "preparer-provided", "ai_confirmed": "AI-proposed · confirmed by human",
               "ai_assumed": "AI-assumed (UNCONFIRMED)"}
_NOT_FINAL = "Draft — not final: contains provisional / unconfirmed lines."


@dataclass
class MfsLine:
    concept_id: str
    label: str
    section: str
    current: "float | None"
    prior: "float | None"
    ai_assumed: bool = False
    new_concept: bool = False          # rendered on a master concept ADDED this run (AI-proposed)
    derived: bool = False              # a deterministically-computed net/subtotal/total (engine arithmetic)
    kind: str = "leaf"                 # leaf | net | subtotal | total
    l1_group: "str | None" = None      # OCI sub-heading bucket ("Items that WILL/WILL NOT be reclassified")
    judgment_only: bool = False        # a Case-B maturity split — management judgment, no independent check
    split_undetermined: bool = False   # a declared current/non-current pair the TB couldn't split (shown unsplit)
    uncorroborated: bool = False       # a face-sheet split INJECTED with no TB anchor (Slice B2b) — uncorroborated


@dataclass
class MfsExport:
    bank: str
    statements: dict = field(default_factory=dict)        # statement -> [MfsLine]
    provenance: list = field(default_factory=list)        # (account, client_label, concept, prov, conf)
    findings: list = field(default_factory=list)          # (account, client_label, reason)
    extensions: list = field(default_factory=list)        # (canonical, statement, section) added this run
    not_final: bool = False
    unit: str = "SAR'000"                                 # presentation unit caption
    period_current: "str | None" = None                  # period-end date of the current column
    period_prior: "str | None" = None                    # period-end date of the prior column
    statement_titles: dict = field(default_factory=dict)  # statement key -> display title (seed-declared)
    note_status: dict = field(default_factory=dict)       # note key -> BUILT|PARTIAL|BLOCKED|RECONCILED|JUDGMENT_*
    judgment_markers: dict = field(default_factory=dict)  # concept_id -> persistent management-judgment caveat
    judgment_kinds: dict = field(default_factory=dict)    # concept_id -> the marker's KIND for the face tag
    #                                                       ('split' | a seed caveat_kind e.g. 'estimated residual')
    note_results: dict = field(default_factory=dict)      # note key -> raw attached result (MovementNote / split dict)
    note_views: list = field(default_factory=list)        # render-ready RenderableNote per DECLARED note (incl. not-generated)
    sign_candidates: list = field(default_factory=list)   # Slice S6a — deterministic balance-failure sign candidates
    disclaimer: str = ""                                  # prominent up-front honesty line (e.g. synthetic-data notice)

    @property
    def notes_complete(self) -> bool:
        """True only if EVERY attached note is BUILT — a PARTIAL/BLOCKED note means the export is not a
        complete breakdown (queryable, distinct from a generic 'has findings')."""
        return all(s == "BUILT" for s in self.note_status.values())


def build_master_fs_export(master_store, stored: dict, bank: str = None,
                           period_current: str = None, period_prior: str = None,
                           unit: str = "SAR'000", note_results: dict = None, gl: dict = None,
                           disclaimer: str = "", notes_attempted: bool = False) -> MfsExport:
    from ai_accountant.master_fs.model import MasterConcept
    from ai_accountant.master_fs.derive import balance_difference, derive_statement, orphan_leaves
    from ai_accountant.master_fs.render import carried_leaf_amounts
    client = stored.get("client", "client")
    ext_ids = set()
    for e in stored.get("extensions", []):                # re-add this run's AI-proposed master extensions
        master_store.add(MasterConcept(
            concept_id=e["concept_id"], statement=e["statement"], l0_section=e["l0_section"],
            l1_group=None, canonical_concept=e["canonical"], label_aliases={}, presence="client_added",
            order=int(e["order"]), provenance=e.get("provenance", "proposed_unconfirmed")))
        ext_ids.add(e["concept_id"])
    ms = ClientMappingStore()
    for m in stored["mappings"]:
        ms.put(client, MappingRecord(account=m["account"], concept_id=m.get("concept_id"),
                                     client_label=m.get("client_label", ""), provenance=m.get("provenance", ""),
                                     confidence=m.get("confidence", ""), flagged_reason=m.get("flagged_reason", ""),
                                     master_id=master_store.master_id))
    cur = {t["account"]: t["current"] for t in stored["tb"] if t.get("current") is not None}
    pri = {t["account"]: t["prior"] for t in stored["tb"] if t.get("prior") is not None}

    # ---- Slice B2b: read the balance-sheet face split (a SECOND source), reconcile it against the TB's lumped
    # total, and apply it as an OVERRIDE LAYER (the TB stays the anchor — never mutated). The reconcile is the
    # firewall: nc+c must equal the TB total in BOTH years or the pair flags (no half-apply). When no face split
    # is supplied/confirmed this collapses to the flag-don't-lump path (split_undetermined) → byte-identical.
    from ai_accountant.master_fs.face_split import reconcile_face_split
    from ai_accountant.master_fs.render import apply_split_override
    cur_leaf_raw = carried_leaf_amounts(master_store, ms, client, cur)      # the RAW (lumped) per-concept anchor
    pri_leaf_raw = carried_leaf_amounts(master_store, ms, client, pri)
    clean_pairs, ambiguous_pairs = master_store.maturity_pairs()
    confirmed_pairs = [tuple(p) for p in stored.get("maturity_pairs_confirmed", [])]
    _face_raw = stored.get("face_split")
    _face_applied = bool(_face_raw) and bool(stored.get("face_split_confirmed"))
    _rec = reconcile_face_split(clean_pairs + confirmed_pairs, cur_leaf_raw, pri_leaf_raw,
                                (_face_raw or {}).get("amounts", {}) if _face_applied else {})
    split_override = _rec["override"]
    split_undetermined = _rec["undetermined"]
    uncorroborated = _rec["uncorroborated"]
    split_findings = []
    for _cid, _kind, _msg in _rec["findings"]:
        _c = master_store.get(_cid)
        split_findings.append((f"face_only:{_cid}" if _kind == "no_anchor" else f"maturity:{_cid}",
                               _c.canonical_concept if _c else _cid, _msg))
    if _face_applied:
        for _u in (_face_raw.get("unmapped") or []):
            split_findings.append(("face_unmapped", _u.get("label", "?"),
                                   f"{_u.get('reason', 'unmapped')} — flagged, not applied"))
    elif _face_raw:                                # read but NOT confirmed → pending: not_final, not applied
        split_findings.append(("face_split:pending", "current/non-current split",
                               "balance-sheet split read but NOT confirmed — confirm to apply; shown unsplit"))

    # ---- Slice S6a: a per-TB, human-confirmed sign correction for a mis-signed contra (e.g. treasury stored
    # negative under a subtracting formula → double-negation). Applied as a component-sign OVERRIDE to the face
    # totals AND the balance-check; the firewall (the balance-check) RE-PROVES below — it clears only at true zero,
    # so a wrong correction stays flagged. Unconfirmed/absent → {} → no change (byte-identical).
    sign_overrides = (stored.get("sign_corrections") or {}) if stored.get("sign_corrections_confirmed") else {}

    mapping = ms.mapping(client)
    ai_assumed_concepts = {r.concept_id for r in mapping.values() if r.provenance == "ai_assumed" and r.concept_id}

    # build the declared notes unless caller pre-built them. Movement notes need a GL; STATIC BREAKDOWNS need
    # none (they re-present the concept's TB leaves), so a notes-attempted TB-only run still builds them.
    if note_results is None and (gl or notes_attempted):
        from ai_accountant.master_fs.notes import build_declared_notes
        acct_levels = {t["account"]: t.get("levels", []) for t in stored["tb"]}   # TB level hierarchy → tiers
        agenda = stored.get("agenda")                        # AI-proposed concept list; None → seed static floor
        note_results = build_declared_notes(master_store, ms, client, cur, gl or {},
                                            breakdowns=stored.get("breakdowns"), account_levels=acct_levels,
                                            prior_amounts=(pri or None),
                                            agenda=(agenda.get("buildable") if isinstance(agenda, dict) else agenda))

    statements = {}
    for st in master_store.statement_keys():                # statement set is seed-declared
        rows = render_comparative(master_store, ms, client, cur, pri, bank, st, split_override=split_override,
                                  sign_overrides=sign_overrides)
        # the LINE is the master concept (leaves: many client accounts roll into it; the client's OWN
        # captions live in the provenance view below). Net/subtotal/total lines are engine-DERIVED.
        statements[st] = [MfsLine(concept_id=r.concept_id, label=r.label, section=r.section,
                                  current=r.current, prior=r.prior,
                                  ai_assumed=(not r.derived) and r.concept_id in ai_assumed_concepts,
                                  new_concept=(not r.derived) and r.concept_id in ext_ids,
                                  derived=r.derived, kind=r.kind, l1_group=r.l1_group) for r in rows]

    provenance, findings, sign_candidates = [], [], []
    for r in sorted(mapping.values(), key=lambda x: x.account):
        if r.concept_id:
            concept = master_store.get(r.concept_id)
            provenance.append((r.account, r.client_label, concept.canonical_concept if concept else r.concept_id,
                               r.provenance, r.confidence))
        else:
            findings.append((r.account, r.client_label, r.flagged_reason or "unmapped — left open"))

    # GUARDS (deterministic, current period): orphaned populated leaves + the seed's balance check. The leaf
    # map is POST-OVERRIDE so the balance check runs on the applied split — a BS-only (no-anchor) injection
    # that unbalances the SOFP is caught here and flagged loudly, never auto-corrected (the branch-3 guardrail).
    cur_leaf = apply_split_override(cur_leaf_raw, {cid: o.get("current") for cid, o in split_override.items()})
    bc = master_store.balance_check()
    for st in master_store.statement_keys():
        st_leaf = {cid: a for cid, a in cur_leaf.items()
                   if master_store.get(cid) and master_store.get(cid).statement == st}
        for orph in orphan_leaves(master_store, st, set(st_leaf)):
            c = master_store.get(orph)
            findings.append((orph, c.canonical_concept if c else orph,
                             "populated leaf not included in any section total — total understated until "
                             "it is attached to a total (the AI-added concept has no total-membership yet)"))
        if bc and st == bc.get("statement"):
            values, _ = derive_statement(master_store, st, st_leaf, sign_overrides=sign_overrides)   # RE-PROOF
            diff = balance_difference(master_store, values)
            if diff is not None and abs(diff) >= 0.005:
                findings.append(("balance check", bc.get("label", "Balance sheet"),
                                 f"{bc.get('short', 'BS')} does not balance (current period): assets "
                                 f"{values[bc['assets_total']]:,.0f} vs liabilities+equity "
                                 f"{values[bc['liab_equity_total']]:,.0f}, difference {diff:,.0f} — render "
                                 f"shown as-is, never auto-corrected"))
                if not sign_overrides:                            # Slice S6a: deterministic diagnosis for the gate
                    from ai_accountant.master_fs.derive import diagnose_balance
                    sign_candidates = diagnose_balance(master_store, st, st_leaf, diff)

    # MATURITY-PAIR guard: the per-pair reconcile ran early (Slice B2b's reconcile_face_split, which subsumes the
    # flag-don't-lump path). It produced `split_undetermined` (no face split / disagreement → shown unsplit),
    # `split_override` (a reconciled or face-only split, applied above), `uncorroborated` (no-anchor injection)
    # and `split_findings`. Surface those findings; an ambiguous pairing the human hasn't resolved is flagged too.
    findings.extend(split_findings)
    confirmed_ids = {i for p in confirmed_pairs for i in p}
    for amb in ambiguous_pairs:
        if not (set(amb.get("nc", []) + amb.get("c", [])) & confirmed_ids):
            findings.append((f"maturity_pair:{amb['base']}", amb["base"],
                             "current/non-current pairing ambiguous (no clean nc/c match) — not paired; "
                             "confirm the pairing before relying on the split"))

    # attached note results (conditional — only when a GL was supplied for a concept that declares a note):
    # a PARTIAL/BLOCKED note's coverage finding propagates to findings, and its status is queryable, so a
    # not-fully-covered note can never present as complete. (None → no change → byte-identical.)
    note_status_map: dict = {}
    judgment_markers: dict = {}
    judgment_kinds: dict = {}                              # concept_id -> the marker KIND (face-tag label)
    if note_results:
        from ai_accountant.master_fs.notes import note_findings, note_status
        for key, r in note_results.items():
            if "anchor" in r:                              # single-concept note (N0/N0.1)
                anchor = r.get("anchor")
                caption = r.get("caption", "note")          # the note's own caption (mechanism-agnostic)
                findings.extend(note_findings(key, anchor, caption))
                findings.extend((f"sign:{key}", caption, s) for s in r.get("sign_findings", []))
                note_status_map[key] = note_status(anchor)
            if "split" in r:                               # one-total → N-concepts split note (N0.2)
                sr = r["split"]
                findings.extend(sr.get("findings", []))
                note_status_map[key] = sr.get("status", "—")
                if sr.get("judgment_marker"):              # PERSISTENT — survives confirmation
                    for c in sr.get("split_concepts", []):
                        judgment_markers[c] = sr["judgment_marker"]
                        judgment_kinds[c] = "split"        # a current/non-current split judgment
    # SEED-DECLARED honesty caveat (SL-1b): a populated leaf carrying a `caveat` in the seed (e.g. an estimated
    # balancing residual) raises a FACE-VISIBLE finding + a persistent judgment marker, and keeps the statement
    # NOT-FINAL — so the figure is shown AND flagged, never a silent clean line. No literal in engine code; the
    # caveat text + which concept carries it are SEED DATA.
    for lines in statements.values():
        for l in lines:
            mc = master_store.get(l.concept_id)
            if mc is not None and getattr(mc, "caveat", "") and not l.derived \
                    and (l.current not in (None, 0.0) or l.prior not in (None, 0.0)):
                judgment_markers[l.concept_id] = mc.caveat
                judgment_kinds[l.concept_id] = getattr(mc, "caveat_kind", "") or "caveat"   # seed-declared kind
                if not any(f[0] == f"caveat:{l.concept_id}" for f in findings):
                    findings.append((f"caveat:{l.concept_id}", l.label, mc.caveat))
    for lines in statements.values():                      # mark the judgment-only / split-undetermined lines
        for l in lines:
            if l.concept_id in judgment_markers:
                l.judgment_only = True
            if l.concept_id in split_undetermined:
                l.split_undetermined = True
            if l.concept_id in uncorroborated:
                l.uncorroborated = True

    # RENDER-READY note views — built ONLY when a notes run was attempted (note_results present). A
    # faces-only run (no GL) yields no note views → byte-identical to today. Within a notes run, a declared
    # note that did NOT build gets an explicit "not generated" view, never silently absent.
    # `notes_attempted` forces the per-note "not generated" cards even with NO GL (the TB-only honest
    # output: every declared roll-forward note renders "not generated — no movement data", never faked).
    face_by_concept = {l.concept_id: l for st in statements.values() for l in st if l.concept_id}
    note_views = (_build_note_views(master_store, note_results or {}, face_by_concept)
                  if (note_results or notes_attempted) else [])
    not_final_notes = any(v.status not in ("BUILT", "RECONCILED", "not generated") for v in note_views)

    extensions = [(e["canonical"], e["statement"], e["l0_section"]) for e in stored.get("extensions", [])]
    not_final = bool(ai_assumed_concepts) or bool(findings) or bool(extensions) or not_final_notes
    titles = {k: master_store.statement_title(k) for k in master_store.statement_keys()}
    return MfsExport(bank=bank, statements=statements, provenance=provenance, findings=findings,
                     extensions=extensions, not_final=not_final, unit=unit,
                     period_current=period_current, period_prior=period_prior, statement_titles=titles,
                     note_status=note_status_map, judgment_markers=judgment_markers, judgment_kinds=judgment_kinds,
                     note_results=note_results or {}, note_views=note_views, sign_candidates=sign_candidates,
                     disclaimer=disclaimer)


def _note_caption(master_store, decl, key):
    c = master_store.get(key)
    return decl.get("caption") or (c.canonical_concept if c and c.canonical_concept else None) or key


def _build_note_views(master_store, note_results, face_by_concept=None) -> list:
    """One RenderableNote per declared note. Built movement note → renderable(); built split → renderable_split();
    declared-but-unbuilt → an explicit 'not generated' view. Status is master-FS (anchor/split), loud in words.
    When a note can't build its breakdown, the view shows the concept's FACE-STATEMENT line (label + amounts)
    verbatim instead of an empty placeholder — `face_by_concept` maps concept_id → its MfsLine."""
    from ai_accountant.reporting.render_model import (RenderableNote, RenderRow, renderable, renderable_split,
                                                      renderable_static, status_caption)
    from ai_accountant.master_fs.notes import note_findings, note_status
    face_by_concept = face_by_concept or {}
    views: list = []
    declared_keys = set()
    for decl in master_store.notes():
        key = decl.get("concept") or decl.get("note")
        if key in declared_keys:                           # DUAL-DECLARED (roll_forward + static_breakdown) →
            continue                                       # render ONCE. note_results[key] already holds the
        declared_keys.add(key)                             # GL-presence pick (movement if GL, else static) — S3b
        caption = _note_caption(master_store, decl, key)
        r = note_results.get(key)
        if r is None:                                      # declared but not built — explicit, never faked
            mech = decl.get("mechanism")
            reason = ("no general ledger supplied — the movement breakdown is not generated" if mech == "roll_forward"
                      else "concept not populated by this trial balance" if mech == "static_breakdown"
                      else f"mechanism '{mech}' not yet built (its own future slice)")
            # show the concept's FACE line(s) (label + figure, as on the statement) rather than an empty
            # placeholder. A note covers either one `concept` or several `splits` (e.g. current/non-current).
            concept_ids = ([key] if decl.get("concept")
                           else [s["concept"] for s in decl.get("splits", []) if s.get("concept")])
            face_rows = []
            for cid in concept_ids:
                fl = face_by_concept.get(cid)
                if fl is not None and (fl.current is not None or fl.prior is not None):
                    hp = fl.prior is not None
                    face_rows.append(RenderRow(fl.label, fl.current, fl.current, kind="line",
                                               prior_value=(fl.prior if hp else None),
                                               prior_raw_sar=(fl.prior if hp else None)))
            if face_rows:                                  # the figure(s) carried to the face — shown verbatim
                cav = status_caption("not generated",
                                     f"{reason}; the amount(s) shown are the figure(s) carried to the face statement")
                views.append(RenderableNote(note_ref=caption, status="not generated", status_line=cav,
                                            caveats=[cav], reasons=[], unit_label="", unit_confirmed=True, rows=face_rows))
            else:                                          # no face figure either → the honest empty placeholder
                loud = status_caption("not generated", reason)
                views.append(RenderableNote(note_ref=caption, status="not generated", status_line=loud,
                                            caveats=[loud], reasons=[], unit_label="", unit_confirmed=False, rows=[]))
        elif "note" in r:                                  # a movement note (any roll-forward)
            rn = renderable(r["note"])
            anchor = r.get("anchor")
            st = note_status(anchor)
            detail = "; ".join(m for *_a, m in note_findings(key, anchor, caption))
            rn.status = st
            rn.status_line = status_caption(st, detail)
            rn.caveats = [status_caption(st, detail)] + list(r.get("sign_findings", [])) + list(rn.caveats)
            views.append(rn)
        elif "split" in r:                                 # one-total → N-concepts split
            sr = r["split"]
            captions = {c: _note_caption(master_store, {}, c) for c in sr.get("split_concepts", [])}
            views.append(renderable_split(sr, captions))
        elif "static" in r:                                # one concept → N component lines (no GL)
            views.append(renderable_static(r["static"]))
        elif "register" in r:                              # SL-1c.1 — a register-enriched N-attribute table
            from ai_accountant.reporting.render_model import renderable_register
            views.append(renderable_register(r["register"]))
    # AGENDA-built notes (Slice S3c): an AI-proposed static note on a concept the seed did NOT declare is in
    # note_results but not in the declared loop — render it too (still foot-checked; never faked).
    for key, r in note_results.items():
        if key not in declared_keys and isinstance(r, dict) and "static" in r:
            views.append(renderable_static(r["static"]))
    return views


# ---- Excel ---------------------------------------------------------------------------------------
_AMBER = PatternFill("solid", fgColor="B26A00")
_RED = PatternFill("solid", fgColor="B00020")
_GREY = PatternFill("solid", fgColor="EEEEEE")
_WHITE_BOLD = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)
_ITAL = Font(italic=True, size=9)


def _num(v):
    return "—" if v is None else round(v, 2)


def _acct(v, decimals: int = 0):
    """Accounting presentation for PDF text: a negative reads as ``(1,234)``, never ``-1,234``; None → em dash.
    (Excel keeps real numbers and gets the same look via a cell number_format, so footing/summability survive.)"""
    if v is None:
        return "—"
    s = f"{abs(v):,.{decimals}f}"
    return f"({s})" if v < 0 else s


def export_master_fs_excel(model: MfsExport, path) -> str:
    wb = Workbook()
    wb.remove(wb.active)
    cur_head = model.period_current or "Current"
    pri_head = model.period_prior or "Prior"
    for st, lines in model.statements.items():
        title = model.statement_titles.get(st, st)
        ws = wb.create_sheet(title[:31])
        r0 = 1
        if model.disclaimer:                                 # prominent up-front honesty (e.g. synthetic data)
            ws.cell(r0, 1, f"⚠ {model.disclaimer}").font = Font(bold=True, color="B26A00")
            r0 += 1
        ws.cell(r0, 1, f"{title} — {(model.bank or 'representative bank').title()} (illustrative)").font = _BOLD
        ws.cell(r0 + 1, 1, f"Expressed in {model.unit}").font = _ITAL
        hdr = r0 + 2
        for col, head in ((1, "Line"), (2, cur_head), (3, pri_head), (4, "Classification")):
            ws.cell(hdr, col, head).font = _BOLD
        ws.freeze_panes = ws.cell(hdr + 1, 1)
        row, section, group = hdr + 1, None, None
        for l in lines:
            if l.section != section:
                section = l.section
                group = None
                ws.cell(row, 1, section).font = _BOLD
                row += 1
            if l.l1_group and l.l1_group != group:               # OCI two-bucket sub-heading
                group = l.l1_group
                ws.cell(row, 1, "  " + l.l1_group).font = _ITAL
                row += 1
            lbl = ws.cell(row, 1, "    " + l.label)
            c2 = ws.cell(row, 2, _num(l.current)); c2.alignment = Alignment(horizontal="right")
            c3 = ws.cell(row, 3, _num(l.prior)); c3.alignment = Alignment(horizontal="right")
            c2.number_format = c3.number_format = "#,##0;(#,##0)"   # bracket-negatives; cells stay numbers (GUARD 2)
            if l.derived:                                        # computed total/subtotal — bold, NEUTRAL
                lbl.font = c2.font = c3.font = _BOLD              # (never amber: it is arithmetic, not an AI line)
            else:
                tag_txt = " · ".join((["NEW concept (AI-proposed)"] if l.new_concept else [])
                                     + (["AI assumption — not settled"] if l.ai_assumed else [])
                                     + (["MANAGEMENT JUDGMENT (split — no independent verification)"]
                                        if l.judgment_only else [])
                                     + (["SPLIT UNDETERMINED — total shown per the TB; breakdown in the notes"]
                                        if l.split_undetermined else [])
                                     + (["BS-ONLY — NOT RECONCILED to the TB (from the balance sheet)"]
                                        if l.uncorroborated else []))
                if tag_txt:
                    tag = ws.cell(row, 4, tag_txt)
                    tag.font, tag.fill = _WHITE_BOLD, _AMBER
            row += 1
        for c, w in (("A", 50), ("B", 18), ("C", 18), ("D", 32)):
            ws.column_dimensions[c].width = w
        if model.not_final:
            ws.cell(row + 1, 1, _NOT_FINAL).font = _ITAL          # small plain footnote, conditional
        ws.cell(row + 2, 1, "Derived totals/subtotals are computed deterministically by the engine; "
                            "AI-mapped leaves are tagged.").font = _ITAL
        if not model.note_views:                             # faces-only doc → keep the honest "no notes" line
            ws.cell(row + 3, 1, "Note-by-note breakdowns are a later phase; this shows the face statements "
                                "only.").font = _ITAL

    # mappings-applied / provenance view — always present
    ws = wb.create_sheet("Mappings applied")
    ws.cell(1, 1, "MAPPINGS APPLIED — each account → its master concept, and WHO decided it").font = _BOLD
    for col, head in ((1, "Account"), (2, "Client label"), (3, "Master concept"), (4, "Provenance"), (5, "Confidence")):
        ws.cell(3, col, head).font = _BOLD
    ws.freeze_panes = ws.cell(4, 1)
    r = 4
    for account, label, concept, prov, conf in model.provenance:
        ws.cell(r, 1, account); ws.cell(r, 2, label); ws.cell(r, 3, concept)
        pc = ws.cell(r, 4, _PROV_LABEL.get(prov, prov))
        if prov == "ai_assumed":
            pc.font, pc.fill = _WHITE_BOLD, _AMBER                # AI-assumed never looks human-confirmed
        ws.cell(r, 5, conf)
        r += 1
    for c, w in (("A", 12), ("B", 42), ("C", 42), ("D", 32), ("E", 12)):
        ws.column_dimensions[c].width = w

    # findings — the unmapped / flagged items, surfaced
    ws = wb.create_sheet("Findings")
    ws.cell(1, 1, "OPEN FINDINGS — accounts NOT placed on a face line (surfaced, not guessed)").font = _BOLD
    for col, head in ((1, "Account"), (2, "Client label"), (3, "Why open")):
        ws.cell(3, col, head).font = _BOLD
    r = 4
    for account, label, reason in (model.findings or [("—", "—", "none — every account mapped")]):
        ws.cell(r, 1, account); ws.cell(r, 2, label); ws.cell(r, 3, reason); r += 1
    for c, w in (("A", 12), ("B", 44), ("C", 50)):
        ws.column_dimensions[c].width = w

    # one sheet per DECLARED note — built notes show their breakdown; declared-but-unbuilt show "not
    # generated". The status colour reinforces the status WORD (which write_note_sheet always writes).
    from ai_accountant.reporting.excel_writer import write_note_sheet
    from ai_accountant.reporting.render_model import status_hex
    for nv in model.note_views:                              # subtle=True: status as a coloured FOOTNOTE word,
        write_note_sheet(wb, nv, status_fill=status_hex(nv.status), subtle=True,   # not a bold banner fill
                         period_current=model.period_current, period_prior=model.period_prior)
    wb.save(path)
    return path


# ---- PDF -----------------------------------------------------------------------------------------
_P_RED = colors.HexColor("#B00020")
_P_AMBER = colors.HexColor("#B26A00")
_P_GREY = colors.HexColor("#EEEEEE")


def _esc(s):
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _footer(canvas, doc, not_final: bool):
    canvas.saveState()
    w, _h = A4
    canvas.setFillColor(colors.grey)
    canvas.setFont("Helvetica-Oblique", 7.5)
    if not_final:
        canvas.drawString(15 * mm, 8 * mm, _NOT_FINAL)            # small plain footnote, not a banner
    canvas.drawRightString(w - 15 * mm, 8 * mm, f"Page {doc.page}")
    canvas.restoreState()


def export_master_fs_pdf(model: MfsExport, path) -> str:
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=13)
    small = ParagraphStyle("small", parent=styles["Italic"], fontSize=8)
    intro = ("Generated against the master FS structure. Per-note breakdowns follow the face statements."
             if model.note_views else
             "Generated against the master FS structure. Note-by-note breakdowns are a later phase — this "
             "shows the face statements only.")
    story = [Paragraph(f"Financial Statements — {(model.bank or 'representative bank').title()} (illustrative figures)", h1)]
    if model.disclaimer:                                          # prominent up-front honesty (synthetic data)
        disc = ParagraphStyle("disc", parent=styles["BodyText"], textColor=_P_AMBER, fontSize=9,
                              spaceAfter=3, borderPadding=3)
        story.append(Paragraph(f"<b>⚠ {_esc(model.disclaimer)}</b>", disc))
    story += [Paragraph(intro, small), Spacer(1, 4 * mm)]

    cur_head = model.period_current or "Current"
    pri_head = model.period_prior or "Prior"
    for idx, (st, lines) in enumerate(model.statements.items()):
        if idx:
            story.append(PageBreak())          # each statement starts on its OWN page — never straddles a
            #                                    boundary (so copied/extracted text never appears scrambled)
        story.append(Paragraph(f"{model.statement_titles.get(st, st)} — expressed in {_esc(model.unit)}", h1))
        data = [["Line", cur_head, pri_head, ""]]
        sty, section, group = [], None, None
        for l in lines:
            if l.section != section:
                section = l.section
                group = None
                data.append([section, "", "", ""])
                sty.append(("FONTNAME", (0, len(data) - 1), (-1, len(data) - 1), "Helvetica-Bold"))
            if l.l1_group and l.l1_group != group:               # OCI two-bucket sub-heading
                group = l.l1_group
                data.append(["  " + l.l1_group, "", "", ""])
                sty.append(("FONTNAME", (0, len(data) - 1), (0, len(data) - 1), "Helvetica-Oblique"))
            i = len(data)
            cur = _acct(l.current, 0)                             # negatives in brackets — (1,234)
            pri = _acct(l.prior, 0)
            if l.derived:                                        # computed total/subtotal — bold, NEUTRAL
                data.append(["    " + l.label, cur, pri, ""])
                sty.append(("FONTNAME", (0, i), (2, i), "Helvetica-Bold"))
                sty.append(("LINEABOVE", (1, i), (2, i), 0.4, _P_GREY))
            else:
                tag_txt = " · ".join((["NEW"] if l.new_concept else [])
                                     + (["AI assumption"] if l.ai_assumed else [])
                                     + (["MGMT JUDGMENT (split)"] if l.judgment_only else [])
                                     + (["SPLIT UNDETERMINED"] if l.split_undetermined else [])
                                     + (["BS-ONLY · NOT RECONCILED"] if l.uncorroborated else []))
                data.append(["    " + l.label, cur, pri, tag_txt])
                if l.ai_assumed or l.new_concept or l.judgment_only or l.split_undetermined or l.uncorroborated:
                    sty.append(("TEXTCOLOR", (3, i), (3, i), _P_AMBER))
                    sty.append(("FONTNAME", (3, i), (3, i), "Helvetica-Bold"))
        t = Table(data, colWidths=[92 * mm, 30 * mm, 30 * mm, 28 * mm])
        t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), _P_GREY), ("ALIGN", (1, 0), (2, -1), "RIGHT"),
                               ("FONTSIZE", (0, 0), (-1, -1), 8.5), *sty]))
        story += [t, Spacer(1, 6 * mm)]

    # one page per DECLARED note — built breakdown OR an explicit "not generated" page (never absent). The notes
    # follow the faces directly; the AI's working (mappings / extensions / findings) is relegated to the END.
    for nv in model.note_views:
        story.append(PageBreak())
        _note_page(story, nv, styles, h1, model.period_current, model.period_prior)

    # ---- the AI's working — proposed extensions, mappings applied, open findings — AFTER the statements + notes
    story.append(PageBreak())
    if model.extensions:
        story.append(Paragraph("Master extensions proposed this run (AI-proposed, unconfirmed — the "
                               "master grew to fit this client): "
                               + "; ".join(f"<b>{_esc(c)}</b> → {_esc(st)}/{_esc(sec)}"
                                           for c, st, sec in model.extensions), small))
    # mappings-applied / provenance. EVERY data cell is a Paragraph (incl. the account) so a long, space-less
    # id wraps inside its column instead of overprinting the next one; widths sum to the ~159mm usable page.
    cell_p = ParagraphStyle("cell_p", parent=styles["BodyText"], fontSize=7.5, leading=9)
    story.append(Paragraph("Mappings applied — account → master concept, and who decided it", h1))
    pdata = [["Account", "Client label", "Master concept", "Provenance", "Conf."]]
    psty = []
    for i, (account, label, concept, prov, conf) in enumerate(model.provenance, start=1):
        pdata.append([Paragraph(_esc(account), cell_p), Paragraph(_esc(label), cell_p),
                      Paragraph(_esc(concept), cell_p),
                      Paragraph(_esc(_PROV_LABEL.get(prov, prov)), cell_p), conf])
        if prov == "ai_assumed":
            psty.append(("TEXTCOLOR", (3, i), (3, i), _P_AMBER))
            psty.append(("FONTNAME", (3, i), (3, i), "Helvetica-Bold"))
    pt = Table(pdata, colWidths=[16 * mm, 42 * mm, 44 * mm, 45 * mm, 12 * mm])
    pt.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), _P_GREY), ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"), ("GRID", (0, 0), (-1, -1), 0.25, _P_GREY), *psty]))
    story += [pt, PageBreak()]

    # findings
    story.append(Paragraph("Open findings — accounts not placed on a face line", h1))
    story.append(Paragraph("Surfaced, not guessed — none silently auto-answered.", small))
    fdata = [["Account", "Client label", "Why open"]]
    for account, label, reason in (model.findings or [("—", "—", "none — every account mapped")]):
        fdata.append([Paragraph(_esc(account), cell_p), Paragraph(_esc(label), cell_p),
                      Paragraph(_esc(reason), cell_p)])     # account is a Paragraph too → wraps, never overprints
    ft = Table(fdata, colWidths=[42 * mm, 42 * mm, 75 * mm])
    ft.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), _P_GREY), ("FONTSIZE", (0, 0), (-1, -1), 8),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"), ("GRID", (0, 0), (-1, -1), 0.25, _P_GREY)]))
    story += [ft]

    target = path if hasattr(path, "write") else str(path)       # path OR a file-like (bytes wrapper)
    doc = SimpleDocTemplate(target, pagesize=A4, bottomMargin=16 * mm)
    foot = partial(_footer, not_final=model.not_final)
    doc.build(story, onFirstPage=foot, onLaterPages=foot)
    return path


def _register_table(story, rn, styles) -> None:
    """SL-1c.1 — emit a register-enriched note as an N-COLUMN table: the register's attribute columns
    (`rn.columns`, VERBATIM values from `RenderRow.cells`) + the amount column(s). Column-LIST driven — works
    for ANY register's columns (the SL-1c.2 reuse). Wide content → a small font + landscape-style narrow cols;
    section/sub-total/total rows span the attribute columns and bold."""
    has_prior = any(r.prior_value is not None for r in rn.rows)
    cell_st = ParagraphStyle("regcell", parent=styles["BodyText"], fontSize=6, leading=7)
    hdr_st = ParagraphStyle("reghdr", parent=cell_st, fontName="Helvetica-Bold")
    attr_cols = list(rn.columns)
    head = [Paragraph(_esc(str(c).replace("\n", " ")), hdr_st) for c in attr_cols] + [Paragraph("Current", hdr_st)]
    if has_prior:
        head.append(Paragraph("Prior", hdr_st))
    data, sty = [head], []
    span_to = len(attr_cols) - 1
    for r in rn.rows:
        bold = r.kind in ("section", "subtotal", "total")
        i = len(data)
        if r.kind in ("section", "subtotal", "total") or not r.cells:   # a label row spanning the attribute columns
            cells = [Paragraph(_esc(r.label), hdr_st)] + ["" for _ in attr_cols[1:]]
            if span_to >= 1:
                sty.append(("SPAN", (0, i), (span_to, i)))
        else:
            cells = [Paragraph(_esc(str(v)), cell_st) for v in (list(r.cells) + [""] * len(attr_cols))[:len(attr_cols)]]
        amt = "" if r.value is None else _acct(r.value, 0)
        row = cells + [amt] + ([("" if r.prior_value is None else _acct(r.prior_value, 0))] if has_prior else [])
        data.append(row)
        if bold:
            sty.append(("FONTNAME", (0, i), (-1, i), "Helvetica-Bold"))
    n_amt = 1 + (1 if has_prior else 0)
    aw = 18 * mm
    attr_w = max(8 * mm, (250 * mm - n_amt * aw) / max(len(attr_cols), 1))   # fill a landscape-ish width
    widths = [attr_w] * len(attr_cols) + [aw] * n_amt
    t = Table(data, colWidths=widths, repeatRows=1)
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), _P_GREY), ("VALIGN", (0, 0), (-1, -1), "TOP"),
                           ("ALIGN", (-n_amt, 0), (-1, -1), "RIGHT"), ("FONTSIZE", (0, 0), (-1, -1), 6),
                           ("GRID", (0, 0), (-1, -1), 0.25, _P_GREY), *sty]))
    story.append(t)


def _note_page(story, rn, styles, h1, period_current=None, period_prior=None) -> None:
    """One note breakdown page — the status reads in WORDS (the honesty), colour only reinforces. TEMPORARY
    demo styling: the status is a SUBTLE coloured footnote line, not a bold banner fill; the status TEXT
    label stays. Mirrors the note-pipeline drill-down so a master-FS note carries identical honesty.
    `period_current`/`period_prior` label the comparative columns with the TB's real years (else Current/Prior)."""
    from ai_accountant.reporting.render_model import status_hex
    col = colors.HexColor("#" + status_hex(rn.status))
    st_style = ParagraphStyle("notest", parent=styles["BodyText"], textColor=col, fontSize=9, leading=12)
    cav_style = ParagraphStyle("notecav", parent=styles["Italic"], textColor=col, fontSize=8, leading=10)
    story.append(Paragraph(f"Note — {_esc(rn.note_ref)}", h1))
    final = rn.status in ("BUILT", "RECONCILED", "not generated")    # not-generated is absent, not unfinished
    story.append(Paragraph(f"<b>STATUS: {_esc(rn.status)}</b>" + ("" if final else "  (NOT FINAL)"), st_style))
    for c in rn.caveats:
        story.append(Paragraph(f"!! {_esc(c)}", cav_style))
    story.append(Spacer(1, 2 * mm))
    if not rn.rows:                                              # nothing on the face either — the status + reason
        return                                                   # above already say it; no redundant sentence
    if getattr(rn, "columns", None):                             # SL-1c.1 — REGISTER TABLE: N attribute columns + amount(s)
        _register_table(story, rn, styles)
        return
    has_prior = any(r.prior_value is not None for r in rn.rows)   # a comparative column (Slice S3b)
    # a real period label (the TB's own column header, e.g. "31 Dec 2024") is used VERBATIM — it already
    # carries the period (and often the unit); the "(unit)" suffix only dresses the generic Current/Prior fallback.
    def _hdr(label, fallback):
        return label if label else (f"{fallback} ({rn.unit_label})" if rn.unit_confirmed else fallback)
    amt_hdr = _hdr(period_current if has_prior else None, "Current" if has_prior else "Amount")
    prior_hdr = _hdr(period_prior, "Prior")
    ndata = [["", amt_hdr] + ([prior_hdr] if has_prior else [])]
    nsty = []
    # the label is a Paragraph so a long line (esp. the "Total — agrees to … on the Statement of …" tie) WRAPS
    # inside its column instead of overprinting the amounts; indent via &nbsp; (a Paragraph collapses plain spaces).
    lbl_base = ParagraphStyle("nlbl", parent=styles["BodyText"], fontSize=8.5, leading=10.5)
    lbl_bold = ParagraphStyle("nlblb", parent=lbl_base, fontName="Helvetica-Bold")
    for r in rn.rows:
        val = "WITHHELD" if r.value is None else _acct(r.value, 2)   # negatives in brackets — (1,234.00)
        is_bold = r.kind in ("section", "subtotal", "total")
        label = Paragraph(("&nbsp;" * (4 * r.indent)) + _esc(r.label), lbl_bold if is_bold else lbl_base)
        row_cells = [label, val]
        if has_prior:                                  # withhold the prior column on the SAME R11 rule
            pv = ("WITHHELD" if not rn.unit_confirmed else
                  ("" if r.prior_value is None else _acct(r.prior_value, 2)))
            row_cells.append(pv)
        ndata.append(row_cells)
        if is_bold:                                    # bold the AMOUNT strings (the label self-bolds via its style)
            nsty.append(("FONTNAME", (1, len(ndata) - 1), (-1, len(ndata) - 1), "Helvetica-Bold"))
    widths = [89 * mm, 35 * mm, 35 * mm] if has_prior else [109 * mm, 50 * mm]   # sums to the ~159mm usable width
    nt = Table(ndata, colWidths=widths)
    nt.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), _P_GREY), ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"), ("FONTSIZE", (0, 0), (-1, -1), 8.5), *nsty]))
    story.append(nt)


def export_master_fs_excel_bytes(model: MfsExport) -> bytes:
    import io
    buf = io.BytesIO()
    export_master_fs_excel(model, buf)
    return buf.getvalue()


def export_master_fs_pdf_bytes(model: MfsExport) -> bytes:
    import io
    buf = io.BytesIO()
    export_master_fs_pdf(model, buf)
    return buf.getvalue()
